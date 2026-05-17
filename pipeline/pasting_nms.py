#!/usr/bin/env python3
"""
Build a formula-ready Card Report workbook and refresh master references.

Usage:
    python pasting_nms.py --config ingest.yml

Inputs come from ingest.yml:
  - prepared Card Report: {output_base}/input/NMS/{week_label}_Card Report.xlsx
  - Fiber computed workbook: {computed_dir}/{week_label}_NRR_Fiber_computed.xlsx
  - All Fiber + OMSP outputs: pipeline.all_fiber_output, pipeline.dwdm_omsp_output
  - master workbook: pipeline.master_workbook

Run after update_all_fiber.py so OMSP library references use the current week's All Fiber data.

Output defaults to:
  {computed_dir}/{week_label}_Card_Report_computed.xlsx
"""

from __future__ import annotations

import importlib.util
from pathlib import Path


def _bootstrap_import_paths() -> None:
    pipeline_dir = next(p for p in Path(__file__).resolve().parents if p.name == "pipeline")
    spec = importlib.util.spec_from_file_location("_pathsetup", pipeline_dir / "_pathsetup.py")
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    module.setup(__file__)


_bootstrap_import_paths()

import argparse
import logging
import re
import shutil
import sys
from pathlib import Path

import openpyxl # type: ignore
import xlwings as xw # type: ignore
from openpyxl import load_workbook # type: ignore

from excel_refs import force_excel_calculate, force_recalc_file
from pipeline_config import (
    build_pipeline_paths,
    default_config_path,
    expand_config_value,
    load_yaml,
    resolve_under,
)

log = logging.getLogger("pasting_nms")

HEADER_ROW = 4
DATA_START = HEADER_ROW + 1
BASE_COLS = 30  # A:AD
OTS_SHEET = "OTS Span Band Based"
HI_SHEET = "hi"
OTS_BEFORE_LOAD_SHEET = "OTS Span Before-Load Based"
OMSP_TEMPLATE_SHEETS = ("TemplateForSystem (2)", "TemplateForSystem")
ATTENUATION_SHEET = "Attenuation"
OAU_SHEET = "OAU"
NMS_HEADER_ROW = 8
NMS_DATA_START = NMS_HEADER_ROW + 1
ATTENUATION_PASTE_START = 7
OAU_SFP_PASTE_START = 6
SFP_DAP_GROUP_PATTERNS = ("DAP", "MD48AFS", "SRAPXF", "OPU")
SFP_MD40_PATTERN = "MD40AFS"

FORMULA_AE = (
    '=LEFT(A{r},4)'
    '&MID(B{r},SEARCH("-",B{r},1)-1,1)'
    '&TEXT(MID(B{r},SEARCH("-",B{r},1)+1,'
    '(SEARCH("-",B{r},SEARCH("-",B{r},1)+1)+1)-SEARCH("-",B{r},1)-2),"00")'
)

FORMULA_AF = (
    '=MID(D{r},'
    'SEARCH("(",D{r},1)+1,'
    '(SEARCH(")",D{r},1)-SEARCH("(",D{r},1))-1)'
)


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def optional_pipeline_path(cfg: dict, config_path: Path, key: str) -> Path | None:
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    raw = pipe.get(key)
    if not isinstance(raw, str) or not raw.strip():
        return None
    return resolve_under(config_path.parent, expand_config_value(raw.strip(), cfg))


def backup_path_for(path: Path, suffix: str) -> Path:
    backup = path.with_stem(path.stem + suffix)
    if not backup.exists():
        return backup
    for i in range(2, 1000):
        candidate = path.with_stem(f"{path.stem}{suffix}_{i}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Could not find available backup filename for {path}")


def make_formula(template: str, row: int) -> str:
    return template.replace("{r}", str(row))


def first_n(values, n: int) -> list[object]:
    row = list(values)[:n]
    if len(row) < n:
        row.extend([None] * (n - len(row)))
    return row


def any_value(values: list[object]) -> bool:
    return any(v is not None for v in values)


def read_optical_attenuation_rows(input_path: Path) -> list[list[object]]:
    """
    Read source Optical Attenuation data as master paste rows:
      source A:G -> master A:G
      source I:J -> master H:I
    """
    if not input_path.is_file():
        raise FileNotFoundError(f"Optical Attenuation input not found: {input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows: list[list[object]] = []
        for row in ws.iter_rows(min_row=NMS_DATA_START, values_only=True):
            vals = first_n(row, 10)
            paste_row = vals[:7] + vals[8:10]
            if any_value(paste_row):
                rows.append(paste_row)
        log.info("Optical Attenuation rows for master: %s", len(rows))
        return rows
    finally:
        wb.close()


def read_sfp_rows(input_path: Path) -> tuple[list[list[object]], list[list[object]]]:
    """
    Read source SFP data as two master paste groups:
      Port contains DAP/MD48AFS/SRAPXF/OPU -> OAU N:O
      Port contains MD40AFS -> OAU U:V
    """
    if not input_path.is_file():
        raise FileNotFoundError(f"SFP input not found: {input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        dap_group: list[list[object]] = []
        md40_group: list[list[object]] = []
        for row in ws.iter_rows(min_row=NMS_DATA_START, min_col=1, max_col=2, values_only=True):
            vals = first_n(row, 2)
            port = str(vals[0] or "").upper()
            if any(pattern in port for pattern in SFP_DAP_GROUP_PATTERNS):
                dap_group.append(vals)
            if SFP_MD40_PATTERN in port:
                md40_group.append(vals)
        log.info("SFP rows for OAU N:O: %s", len(dap_group))
        log.info("SFP rows for OAU U:V: %s", len(md40_group))
        return dap_group, md40_group
    finally:
        wb.close()


def process_card_report(input_path: Path, output_path: Path) -> str:
    if not input_path.is_file():
        raise FileNotFoundError(f"Card Report input not found: {input_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb_in = load_workbook(input_path, read_only=True, data_only=False)
    try:
        ws_in = wb_in.active
        sheet_name = ws_in.title
        log.info("Card input : %s", input_path)
        log.info("Card output: %s", output_path)
        log.info("Card sheet : %s", sheet_name)

        wb_out = openpyxl.Workbook(write_only=True)
        ws_out = wb_out.create_sheet(sheet_name)

        data_rows = 0
        for row_num, row in enumerate(ws_in.iter_rows(values_only=True), start=1):
            vals = first_n(row, BASE_COLS)
            if row_num == HEADER_ROW:
                ws_out.append(vals + ["Formula 1", "Formula 2"])
            elif row_num >= DATA_START:
                ws_out.append(
                    vals
                    + [
                        make_formula(FORMULA_AE, row_num),
                        make_formula(FORMULA_AF, row_num),
                    ]
                )
                data_rows += 1
            else:
                ws_out.append(vals)

        wb_out.save(output_path)
        log.info("Card output saved: %s data rows", data_rows)
        return sheet_name
    finally:
        wb_in.close()


def external_ref(path: Path, sheet_name: str) -> str:
    folder = str(path.parent)
    filename = path.name
    safe_sheet = sheet_name.replace("'", "''")
    return f"'{folder}\\[{filename}]{safe_sheet}'!"


def formula_e(row: int, nrr_ref: str) -> str:
    return f"=IFERROR(INDEX({nrr_ref}$AK:$AK,MATCH(V{row},{nrr_ref}$AV:$AV,0)),D{row})"


def formula_fiber_loss(row: int, nrr_ref: str, match_col: str, fallback_col: str) -> str:
    return (
        f"=IFERROR(INDEX({nrr_ref}$AK:$AK,"
        f"MATCH({match_col}{row},{nrr_ref}$AV:$AV,0)),{fallback_col}{row})"
    )


def formula_fiber_type(row: int, nrr_ref: str, match_col: str) -> str:
    return f"=INDEX({nrr_ref}$K:$K,MATCH({match_col}{row},{nrr_ref}$AW:$AW,0))"


def formula_omsp_d(row: int, nrr_ref: str) -> str:
    return (
        f"=INDEX({nrr_ref}$AV:$AV,MATCH(Q{row},{nrr_ref}$D:$D,0))"
        f'&";"&INDEX({nrr_ref}$AR:$AR,MATCH(Q{row},{nrr_ref}$D:$D,0))'
        f'&";"&INDEX({nrr_ref}$AS:$AS,MATCH(Q{row},{nrr_ref}$D:$D,0))'
    )


def formula_omsp_v(row: int, all_fiber_ref: str) -> str:
    return f"=INDEX({all_fiber_ref}$D:$D,MATCH(S{row},{all_fiber_ref}$B:$B,0))"


def formula_omsp_w(row: int, all_fiber_ref: str) -> str:
    return f"=INDEX({all_fiber_ref}$D:$D,MATCH(U{row},{all_fiber_ref}$B:$B,0))"


def formula_i(row: int, card_ref: str) -> str:
    return (
        f'=IFERROR(INDEX({card_ref}$AF:$AF,MATCH(LEFT(R{row},7),{card_ref}$AE:$AE,0))'
        f'&IF(ISNUMBER(SEARCH("DAP",H{row},1))," - "&INDEX(OAU!O:O,MATCH(\'{OTS_SHEET}\'!R{row},OAU!R:R,0)),'
        f'IF(ISNUMBER(SEARCH("MD40",H{row},1))," - "&INDEX(OAU!V:V,MATCH(LEFT(\'{OTS_SHEET}\'!R{row},7),OAU!Y:Y,0)),'
        f'IF(ISNUMBER(SEARCH("RAPXF",H{row},1))," - "&INDEX(OAU!O:O,MATCH(\'{OTS_SHEET}\'!R{row},OAU!R:R,0)),""))),"Not Found!")'
    )


def formula_m(row: int, card_ref: str) -> str:
    return (
        f'=IFERROR(INDEX({card_ref}$AF:$AF,MATCH(LEFT(T{row},7),{card_ref}$AE:$AE,0))'
        f'&IF(ISNUMBER(SEARCH("DAP",L{row},1))," - "&INDEX(OAU!O:O,MATCH(\'{OTS_SHEET}\'!T{row},OAU!R:R,0)),'
        f'IF(ISNUMBER(SEARCH("MD40",L{row},1))," - "&INDEX(OAU!V:V,MATCH(LEFT(\'{OTS_SHEET}\'!T{row},7),OAU!Y:Y,0)),'
        f'IF(ISNUMBER(SEARCH("RAPXF",L{row},1))," - "&INDEX(OAU!O:O,MATCH(\'{OTS_SHEET}\'!T{row},OAU!R:R,0)),""))),"Not Found!")'
    )


def formula_p(row: int, card_ref: str) -> str:
    return (
        f'=IFERROR(INDEX({card_ref}$AF:$AF,MATCH(LEFT(U{row},7),{card_ref}$AE:$AE,0))'
        f'&IF(ISNUMBER(SEARCH("DAP",O{row},1))," - "&INDEX(OAU!O:O,MATCH(\'{OTS_SHEET}\'!T{row},OAU!R:R,0)),'
        f'IF(ISNUMBER(SEARCH("MD40",O{row},1))," - "&INDEX(OAU!V:V,MATCH(LEFT(\'{OTS_SHEET}\'!T{row},7),OAU!Y:Y,0)),"")),"")'
    )


def fill_formula_column(ws: xw.Sheet, col: str, last_row: int, formula: str) -> None:
    ws.range(f"{col}3").formula = formula
    if last_row > 3:
        ws.range(f"{col}3:{col}{last_row}").api.FillDown()


def fill_formula_column_from(ws: xw.Sheet, col: str, start_row: int, last_row: int, formula: str) -> None:
    ws.range(f"{col}{start_row}").formula = formula
    if last_row > start_row:
        ws.range(f"{col}{start_row}:{col}{last_row}").api.FillDown()


def refresh_fiber_reference_formulas(wb: xw.Book, nrr_ref: str) -> None:
    """
    Refresh known Fiber/NRR external-reference formulas in master template sheets.
    This replaces stale [n]Fiber links with the current NRR workbook path.
    """
    specs = [
        (OTS_SHEET, [("E", formula_fiber_loss(3, nrr_ref, "V", "D")), ("BC", formula_fiber_type(3, nrr_ref, "BH"))]),
        (HI_SHEET, [("J", formula_fiber_loss(3, nrr_ref, "AA", "I")), ("BQ", formula_fiber_type(3, nrr_ref, "CJ"))]),
        (
            OTS_BEFORE_LOAD_SHEET,
            [("J", formula_fiber_loss(3, nrr_ref, "AA", "I")), ("BQ", formula_fiber_type(3, nrr_ref, "CJ"))],
        ),
    ]

    for sheet_name, formulas in specs:
        try:
            ws = wb.sheets[sheet_name]
        except Exception:
            log.warning("  Sheet missing for Fiber refresh: %s", sheet_name)
            continue
        last_row = max(ws.used_range.last_cell.row, 3)
        for col, formula in formulas:
            fill_formula_column(ws, col, last_row, formula)
        log.info("  Fiber/NRR refs refreshed on %s rows 3:%s", sheet_name, last_row)


def refresh_omsp_references(
    omsp_path: Path,
    nrr_path: Path,
    all_fiber_path: Path,
    make_backup: bool,
) -> bool:
    if not omsp_path.is_file():
        log.warning("OMSP workbook not found, skipping reference refresh: %s", omsp_path)
        return True
    if not all_fiber_path.is_file():
        log.error("All Fiber workbook not found for OMSP V/W refresh: %s", all_fiber_path)
        return False

    if make_backup:
        backup = backup_path_for(omsp_path, "_BACKUP")
        shutil.copy2(omsp_path, backup)
        log.info("OMSP backup -> %s", backup.name)

    nrr_ref = external_ref(nrr_path, "Fiber")
    all_fiber_ref = external_ref(all_fiber_path, "Library FIU and OSC Connection")

    log.info("Refreshing OMSP references: %s", omsp_path)
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = None
    try:
        wb = app.books.open(str(omsp_path), update_links=False)
        for sheet_name in OMSP_TEMPLATE_SHEETS:
            try:
                ws = wb.sheets[sheet_name]
            except Exception:
                continue
            last_row = max(ws.used_range.last_cell.row, 9)
            fill_formula_column_from(ws, "D", 9, last_row, formula_omsp_d(9, nrr_ref))
            fill_formula_column_from(ws, "V", 9, last_row, formula_omsp_v(9, all_fiber_ref))
            fill_formula_column_from(ws, "W", 9, last_row, formula_omsp_w(9, all_fiber_ref))
            log.info("  OMSP refs refreshed on %s rows 9:%s", sheet_name, last_row)

        app.calculate()
        wb.save()
        log.info("OMSP workbook saved.")
        return True
    except Exception as exc:
        log.error("OMSP reference refresh failed for %s: %s", omsp_path.name, exc)
        log.error("Manual fix: open the OMSP workbook in Excel, then run pasting_nms.py again.")
        return False
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        try:
            app.screen_updating = True
        except Exception:
            pass
        try:
            app.quit()
        except Exception:
            pass


def paste_block(sheet: xw.Sheet, start_cell: str, rows: list[list[object]]) -> None:
    if not rows:
        return
    start = sheet.range(start_cell)
    sheet.range((start.row, start.column)).value = rows


def paste_nms_sources(
    wb: xw.Book,
    attenuation_rows: list[list[object]],
    sfp_dap_rows: list[list[object]],
    sfp_md40_rows: list[list[object]],
) -> None:
    log.info("Pasting NMS data into master sheets ...")

    ws_att = wb.sheets[ATTENUATION_SHEET]
    att_last = max(ws_att.used_range.last_cell.row, ATTENUATION_PASTE_START)
    ws_att.range(f"A{ATTENUATION_PASTE_START}:I{att_last}").clear_contents()
    paste_block(ws_att, f"A{ATTENUATION_PASTE_START}", attenuation_rows)
    log.info("  Attenuation A:I <- %s rows", len(attenuation_rows))

    ws_oau = wb.sheets[OAU_SHEET]
    oau_last = max(ws_oau.used_range.last_cell.row, OAU_SFP_PASTE_START)
    ws_oau.range(f"N{OAU_SFP_PASTE_START}:O{oau_last}").clear_contents()
    ws_oau.range(f"U{OAU_SFP_PASTE_START}:V{oau_last}").clear_contents()
    paste_block(ws_oau, f"N{OAU_SFP_PASTE_START}", sfp_dap_rows)
    paste_block(ws_oau, f"U{OAU_SFP_PASTE_START}", sfp_md40_rows)
    log.info("  OAU N:O <- %s SFP rows", len(sfp_dap_rows))
    log.info("  OAU U:V <- %s SFP rows", len(sfp_md40_rows))


def refresh_master_references(
    master_path: Path,
    nrr_path: Path,
    card_path: Path,
    card_sheet: str,
    attenuation_rows: list[list[object]],
    sfp_dap_rows: list[list[object]],
    sfp_md40_rows: list[list[object]],
    make_backup: bool,
) -> None:
    if make_backup:
        backup = backup_path_for(master_path, "_BACKUP")
        shutil.copy2(master_path, backup)
        log.info("Master backup -> %s", backup.name)

    nrr_ref = external_ref(nrr_path, "Fiber")
    card_ref = external_ref(card_path, card_sheet)

    log.info("Refreshing master references in %s", OTS_SHEET)
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = None
    try:
        wb = app.books.open(str(master_path), update_links=False)
        paste_nms_sources(wb, attenuation_rows, sfp_dap_rows, sfp_md40_rows)

        refresh_fiber_reference_formulas(wb, nrr_ref)

        ws = wb.sheets[OTS_SHEET]
        last_row = max(ws.used_range.last_cell.row, 3)
        log.info("  Card formula rows on %s: 3:%s", OTS_SHEET, last_row)

        fill_formula_column(ws, "I", last_row, formula_i(3, card_ref))
        fill_formula_column(ws, "M", last_row, formula_m(3, card_ref))
        fill_formula_column(ws, "P", last_row, formula_p(3, card_ref))

        log.info("Force recalculating master after card-report formulas ...")
        force_excel_calculate(app, wb)
        wb.save()
        log.info("Master saved.")
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        app.screen_updating = True
        app.quit()


def run(config_path: Path) -> int:
    cfg = load_yaml(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}

    card_input = paths.nms_week_dir / f"{paths.week_label}_Card Report.xlsx"
    attenuation_input = paths.nms_week_dir / f"{paths.week_label}_Optical_Attenuation.xlsx"
    sfp_input = paths.nms_week_dir / f"{paths.week_label}_SFP.xlsx"
    card_output = optional_pipeline_path(cfg, config_path, "card_report_computed")
    if card_output is None:
        card_output = paths.computed_dir / f"{paths.week_label}_Card_Report_computed.xlsx"
    omsp_path = optional_pipeline_path(cfg, config_path, "dwdm_omsp_output")
    if omsp_path is None:
        omsp_path = paths.output_base / f"{paths.week_label}_OMSP_DWDMEvaluation.xlsx"
    all_fiber_path = optional_pipeline_path(cfg, config_path, "all_fiber_output")
    if all_fiber_path is None:
        all_fiber_path = paths.output_base / f"{paths.week_label}_All Fiber.xlsx"

    sheet_name = process_card_report(card_input, card_output)
    log.info("Recalculating card report computed workbook ...")
    force_recalc_file(card_output)
    attenuation_rows = read_optical_attenuation_rows(attenuation_input)
    sfp_dap_rows, sfp_md40_rows = read_sfp_rows(sfp_input)
    refresh_master_references(
        paths.master_workbook,
        paths.fiber_computed_out,
        card_output,
        sheet_name,
        attenuation_rows,
        sfp_dap_rows,
        sfp_md40_rows,
        make_backup=pipe.get("make_master_backup", True),
    )
    if not refresh_omsp_references(
        omsp_path,
        paths.fiber_computed_out,
        all_fiber_path,
        make_backup=pipe.get("make_master_backup", True),
    ):
        return 1
    return 0


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    parser = argparse.ArgumentParser(description="Build Card Report and refresh master references")
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help=f"Path to ingest YAML (default: {default_config_path()})",
    )
    args = parser.parse_args(argv)
    cfg_path = (args.config or default_config_path()).resolve()
    if not cfg_path.is_file():
        log.error("Config not found: %s", cfg_path)
        return 1
    try:
        return run(cfg_path)
    except Exception:
        log.exception("pasting_nms failed")
        return 1


if __name__ == "__main__":
    sys.exit(main())
