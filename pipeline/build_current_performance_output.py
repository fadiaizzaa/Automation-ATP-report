#!/usr/bin/env python3
"""
Build the merged, formula-ready Current Performance Data workbook.

This is a separate/manual helper. It does not run from run_week_pipeline.py.

Example:
    python build_current_performance_output.py --config ingest.yml --template "C:\\Users\\fadia\\Downloads\\Week 19_Current_Performance_Data_2026-05-07_16-34-38.xlsx"
"""

from __future__ import annotations

import importlib.util
from pathlib import Path


def _bootstrap_import_paths() -> None:
    pipeline_dir = next(p for p in Path(__file__).resolve().parents if p.name == "pipeline")
    spec = importlib.util.spec_from_file_location("_pathsetup", pipeline_dir / "_pathsetup.py")
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    module.setup(__file__)


_bootstrap_import_paths()

import argparse
import logging
import re
import shutil
import sys
import tempfile
import zipfile
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.worksheet.formula import ArrayFormula

from excel_refs import (
    excel_session,
    external_ref,
    fill_formula_column,
    force_excel_calculate,
    restore_workbook_from_backup,
)
from pipeline_config import build_pipeline_paths, expand_config_value, load_yaml, resolve_under


log = logging.getLogger("current_performance_builder")

SHEET1 = "Sheet1"
SHEET2 = "Sheet2"
SHEET3 = "Sheet3"

RAW_COLS = 8
OUTPUT_COLS = 11
MAIN_HEADER_ROW = 8
FIRST_DATA_ROW = 9
CONTINUATION_FIRST_DATA_ROW = 2

FILTER_HEADERS = {
    9: "Filter Line Board",
    10: "Filter Amp Board",
    11: "Filter OSC Board",
}

FILTER_LINE_PATTERNS = [
    "*LOG*",
    "*N30*",
    "*N40*",
    "*N50*",
    "*N60*",
    "*NS4*",
    "*NS3*",
    "*ND2*",
    "*LSX*",
    "*LDX*",
    "*ELOM*",
    "*LSC*",
    "*LTX*",
    "*LQM*",
    "*LDC*",
]
FILTER_AMP_PATTERNS = [
    "*VA*",
    "*OAU*",
    "*OBU*",
    "*RAU*",
    "*RPC*",
    "*DAP*",
    "*MD40*",
    "*RAPXF*",
    "*MR4*",
    "*AFS*",
    "*OPU*",
    "*WSMD*",
]
FILTER_OSC_PATTERNS = ["*ST*", "*SC*"]

OCH_PERFORMANCE_SHEET = "OCH Performance"
OAU_SHEET = "OAU"
OSC_SHEET = "OSC"
OMSP_TEMPLATE_SHEET = "TemplateForSystem (2)"
MASTER_OMSP_BU_SHEET = "OTS Span Band Based"
MASTER_OMSP_BU_COL = "BU"
MASTER_OMSP_BU_MATCH_COL = "BG"
MASTER_OMSP_BU_START_ROW = 3
OCH_PASTE_START_ROW = 7
OAU_PASTE_START_ROW = 6
OSC_PASTE_START_ROW = 3
PASTE_CHUNK_ROWS = 20_000

OCH_PERFORMANCE_EVENTS = {
    "LSIOPCUR",
    "LSOOPCUR",
    "FEC_BEF_COR_ER",
    "FEC_AFT_COR_ER",
    "FEC_BEF_CORER_FLOAT",
    "FEC_AFT_CORER_FLOAT",
    "TDCCUR",
    "DGDCUR",
}
OAU_PERFORMANCE_EVENTS = {"SUMIOPCUR", "SUMOOPCUR"}
OSC_PERFORMANCE_EVENTS = {"LSOOPCUR(DBM)", "LSIOPCUR(DBM)"}


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def excel_array(patterns: list[str]) -> str:
    return "{" + ",".join(f'"{pattern}"' for pattern in patterns) + "}"


def filter_formula(row_idx: int, patterns: list[str]) -> str:
    return f'=SUM(COUNTIF(A{row_idx}, {excel_array(patterns)})) > 0'


def filter_formulas(row_idx: int) -> list[str]:
    return [
        ArrayFormula(f"I{row_idx}", filter_formula(row_idx, FILTER_LINE_PATTERNS)),
        filter_formula(row_idx, FILTER_AMP_PATTERNS),
        filter_formula(row_idx, FILTER_OSC_PATTERNS),
    ]


def event_code(value: object) -> str:
    text = str(value or "").strip().upper()
    return text.split("(", 1)[0].strip()


def event_text(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def pattern_match(value: object, patterns: list[str]) -> bool:
    text = str(value or "").upper()
    return any(pattern.strip("*").upper() in text for pattern in patterns)


def is_och_performance_row(values: list[object]) -> bool:
    return pattern_match(values[0] if values else None, FILTER_LINE_PATTERNS) and event_code(
        values[1] if len(values) > 1 else None
    ) in OCH_PERFORMANCE_EVENTS


def is_oau_row(values: list[object]) -> bool:
    return pattern_match(values[0] if values else None, FILTER_AMP_PATTERNS) and event_code(
        values[1] if len(values) > 1 else None
    ) in OAU_PERFORMANCE_EVENTS


def is_osc_row(values: list[object]) -> bool:
    return pattern_match(values[0] if values else None, FILTER_OSC_PATTERNS) and event_text(
        values[1] if len(values) > 1 else None
    ) in OSC_PERFORMANCE_EVENTS


def copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def styled_row(ws, row_values: list[object], style_cells=None):
    cells = []
    for col_idx, value in enumerate(row_values, start=1):
        cell = WriteOnlyCell(ws, value=value)
        if style_cells and col_idx <= len(style_cells):
            copy_cell_style(style_cells[col_idx - 1], cell)
        cells.append(cell)
    return cells


def first_n(values: Iterable[object], n: int) -> list[object]:
    out = list(values)[:n]
    if len(out) < n:
        out.extend([None] * (n - len(out)))
    return out


def as_master_text(value: object) -> object:
    """Keep performance pasted values as text so master formulas behave like manual paste."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value)


def row_as_master_text(row: list[object]) -> list[object]:
    return [as_master_text(v) for v in first_n(row, RAW_COLS)]


def total_records_text(record_count: int) -> str:
    return f"Total {record_count:,} Records"


def output_default(config_path: Path, week_label: str) -> Path:
    return config_path.parent / "outputs" / f"{week_label}_Current_Performance_Data_output.xlsx"


def is_performance_variant(stem: str) -> bool:
    s = stem.lower()
    return "@1" in s or s.endswith("_1")


def safe_extract_zip(zip_path: Path, dest: Path) -> None:
    dest = dest.resolve()
    with zipfile.ZipFile(zip_path) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            target = (dest / info.filename).resolve()
            try:
                target.relative_to(dest)
            except ValueError as e:
                raise RuntimeError(f"Unsafe path in archive {zip_path.name}: {info.filename!r}") from e
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info, "r") as src, target.open("wb") as out:
                shutil.copyfileobj(src, out)


def list_workbooks(root: Path, suffixes: tuple[str, ...]) -> list[Path]:
    out: list[Path] = []
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in suffixes and not path.name.startswith("~$"):
            out.append(path)
    return sorted(out, key=lambda p: str(p).lower())


def extract_performance_xlsx_from_zip(zip_path: Path, dest: Path) -> list[Path]:
    safe_extract_zip(zip_path, dest)
    xs = list_workbooks(dest, (".xlsx",))
    if not xs:
        raise RuntimeError(f"Performance ZIP has no .xlsx: {zip_path.name}")
    return sorted(xs, key=lambda p: (is_performance_variant(p.stem), p.name.lower()))


def performance_daily_dir(cfg: dict, config_path: Path) -> Path:
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    raw = pipe.get("performance_daily_dir")
    if isinstance(raw, str) and raw.strip():
        return resolve_under(config_path.parent, expand_config_value(raw.strip(), cfg))
    paths = build_pipeline_paths(cfg, config_path)
    return paths.output_base / "input" / "raw files" / "performance_daily"


def sync_performance_workbooks_to_nms(xlsx_files: list[Path], nms_dir: Path, week_label: str) -> None:
    nms_dir.mkdir(parents=True, exist_ok=True)
    for idx, src in enumerate(xlsx_files):
        suffix = "" if idx == 0 else f"_{idx}"
        dest = nms_dir / f"{week_label}_Current Performance Data{suffix}.xlsx"
        shutil.copy2(src, dest)
        log.info("Synced performance workbook -> %s", dest.name)


def archive_processed_daily_zips(daily_dir: Path, zips: list[Path]) -> None:
    if not zips:
        return
    processed_dir = daily_dir / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)
    for zpath in zips:
        target = processed_dir / zpath.name
        if target.exists():
            stem = zpath.stem
            for i in range(2, 1000):
                candidate = processed_dir / f"{stem}_{i}{zpath.suffix}"
                if not candidate.exists():
                    target = candidate
                    break
        shutil.move(str(zpath), str(target))
        log.info("Archived daily ZIP -> %s", target.name)


def ingest_daily_performance_zips(
    cfg: dict,
    config_path: Path,
    *,
    archive: bool = True,
) -> list[Path] | None:
    """
    Extract the newest ZIP in performance_daily_dir (one main + optional _1 pair).
    Returns ordered .xlsx paths, or None if the folder has no ZIPs.
    """
    daily_dir = performance_daily_dir(cfg, config_path)
    if not daily_dir.is_dir():
        return None

    zips = sorted(daily_dir.glob("*.zip"), key=lambda p: p.stat().st_mtime)
    if not zips:
        return None

    latest_zip = zips[-1]
    if len(zips) > 1:
        log.warning(
            "%s daily performance ZIP(s) found; using newest only: %s",
            len(zips),
            latest_zip.name,
        )

    paths = build_pipeline_paths(cfg, config_path)
    with tempfile.TemporaryDirectory(prefix="perf_daily_") as tmp:
        collected = extract_performance_xlsx_from_zip(latest_zip, Path(tmp))

    if not collected:
        raise RuntimeError(f"No performance workbooks extracted from {latest_zip.name}")

    if len(collected) != 2:
        log.warning(
            "Expected 2 performance workbooks in %s, got %s; using first two if available.",
            latest_zip.name,
            len(collected),
        )
        collected = collected[:2]

    sync_performance_workbooks_to_nms(collected, paths.nms_week_dir, paths.week_label)
    log.info(
        "Ingested daily performance ZIP %s -> %s workbook(s) under NMS",
        latest_zip.name,
        len(collected),
    )
    if archive:
        archive_processed_daily_zips(daily_dir, zips)
    return collected


def prepared_performance_paths(cfg: dict, config_path: Path) -> list[Path]:
    paths = build_pipeline_paths(cfg, config_path)
    main = paths.nms_week_dir / f"{paths.week_label}_Current Performance Data.xlsx"
    if not main.is_file():
        raise FileNotFoundError(f"Missing main Current Performance workbook: {main}")

    pattern = f"{paths.week_label}_Current Performance Data_*.xlsx"
    continuations: list[tuple[int, Path]] = []
    for path in paths.nms_week_dir.glob(pattern):
        match = re.fullmatch(
            rf"{re.escape(paths.week_label)}_Current Performance Data_(\d+)\.xlsx",
            path.name,
            flags=re.IGNORECASE,
        )
        if match:
            continuations.append((int(match.group(1)), path))

    return [main] + [path for _, path in sorted(continuations)]


def resolve_performance_files(
    cfg: dict,
    config_path: Path,
    *,
    use_daily: bool = True,
    archive_daily: bool = True,
) -> list[Path]:
    if use_daily:
        daily_files = ingest_daily_performance_zips(cfg, config_path, archive=archive_daily)
        if daily_files:
            return daily_files
    return prepared_performance_paths(cfg, config_path)


def performance_paths(config_path: Path) -> tuple[dict, list[Path]]:
    cfg = load_yaml(config_path)
    return cfg, resolve_performance_files(cfg, config_path)


def optional_pipeline_path(cfg: dict, config_path: Path, key: str) -> Path | None:
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    raw = pipe.get(key)
    if not isinstance(raw, str) or not raw.strip():
        return None
    return resolve_under(config_path.parent, expand_config_value(raw.strip(), cfg))


def backup_path_for(path: Path, suffix: str) -> Path:
    backup = path.with_stem(path.stem + suffix)
    if not backup.exists():
        return backup
    for i in range(2, 1000):
        candidate = path.with_stem(f"{path.stem}{suffix}_{i}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Could not find available backup filename for {path}")


def read_template_styles(template_path: Path):
    wb = load_workbook(template_path, read_only=False, data_only=False)
    try:
        if SHEET1 not in wb.sheetnames:
            raise ValueError(f"Template workbook is missing {SHEET1!r}")
        ws = wb[SHEET1]
        header_styles = {
            row_idx: [ws.cell(row_idx, col_idx) for col_idx in range(1, OUTPUT_COLS + 1)]
            for row_idx in range(1, MAIN_HEADER_ROW + 1)
        }
        data_styles = [ws.cell(FIRST_DATA_ROW, col_idx) for col_idx in range(1, OUTPUT_COLS + 1)]
        widths = {
            letter: ws.column_dimensions[letter].width
            for letter in "ABCDEFGHIJK"
            if ws.column_dimensions[letter].width
        }
        return header_styles, data_styles, widths
    finally:
        wb.close()


def build_output(
    performance_files: list[Path],
    template_path: Path | None,
    output_path: Path,
) -> tuple[list[list[object]], list[list[object]], list[list[object]]]:
    if not performance_files:
        raise ValueError("No Current Performance workbooks were provided")
    for path in performance_files:
        if not path.is_file():
            raise FileNotFoundError(f"Missing Current Performance workbook: {path}")
    style_source_path = template_path or performance_files[0]
    if not style_source_path.is_file():
        raise FileNotFoundError(f"Missing style source workbook: {style_source_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    header_styles, data_styles, widths = read_template_styles(style_source_path)

    workbooks = [load_workbook(path, read_only=True, data_only=False) for path in performance_files]
    try:
        worksheets = [wb[SHEET1] for wb in workbooks]
        main_ws = worksheets[0]
        och_rows: list[list[object]] = []
        oau_rows: list[list[object]] = []
        osc_rows: list[list[object]] = []

        final_last_row = main_ws.max_row + sum(max(0, ws.max_row - 1) for ws in worksheets[1:])
        final_records = final_last_row - (FIRST_DATA_ROW - 1)
        for path, ws in zip(performance_files, worksheets):
            log.info("%s rows: %s", path.name, ws.max_row)
        log.info("Final rows: %s", final_last_row)
        log.info("Final records: %s", final_records)

        out_wb = Workbook(write_only=True)
        out_ws = out_wb.create_sheet(SHEET1)
        for col, width in widths.items():
            out_ws.column_dimensions[col].width = width

        for row_idx, row in enumerate(main_ws.iter_rows(values_only=True), start=1):
            values = first_n(row, RAW_COLS)
            if row_idx == 6:
                values[0] = total_records_text(final_records)
            if row_idx < MAIN_HEADER_ROW:
                values.extend([None, None, None])
            elif row_idx == MAIN_HEADER_ROW:
                values.extend([FILTER_HEADERS[9], FILTER_HEADERS[10], FILTER_HEADERS[11]])
            else:
                values.extend(filter_formulas(row_idx))
                raw_values = row_as_master_text(first_n(row, RAW_COLS))
                if is_och_performance_row(raw_values):
                    och_rows.append(raw_values)
                if is_oau_row(raw_values):
                    oau_rows.append(raw_values)
                if is_osc_row(raw_values):
                    osc_rows.append(first_n(raw_values, 7))

            styles = header_styles.get(row_idx, data_styles)
            out_ws.append(styled_row(out_ws, values, styles))

        next_output_row = main_ws.max_row + 1
        for continuation_ws in worksheets[1:]:
            for source_row_idx, row in enumerate(continuation_ws.iter_rows(values_only=True), start=1):
                if source_row_idx < CONTINUATION_FIRST_DATA_ROW:
                    continue
                values = first_n(row, RAW_COLS)
                text_values = row_as_master_text(values)
                if is_och_performance_row(text_values):
                    och_rows.append(text_values)
                if is_oau_row(text_values):
                    oau_rows.append(text_values)
                if is_osc_row(text_values):
                    osc_rows.append(first_n(text_values, 7))
                values.extend(filter_formulas(next_output_row))
                out_ws.append(styled_row(out_ws, values, data_styles))
                next_output_row += 1

        out_wb.create_sheet(SHEET2)
        out_wb.create_sheet(SHEET3)
        out_wb.save(output_path)
        log.info("Saved -> %s", output_path)
        log.info("OCH Performance filtered rows: %s", len(och_rows))
        log.info("OAU filtered rows: %s", len(oau_rows))
        log.info("OSC filtered rows: %s", len(osc_rows))
        return och_rows, oau_rows, osc_rows
    finally:
        for wb in workbooks:
            wb.close()


def collect_filtered_rows(performance_files: list[Path]) -> tuple[list[list[object]], list[list[object]], list[list[object]]]:
    """Fast path: read raw A:H rows and collect only rows needed by the master."""
    if not performance_files:
        raise ValueError("No Current Performance workbooks were provided")
    for path in performance_files:
        if not path.is_file():
            raise FileNotFoundError(f"Missing Current Performance workbook: {path}")

    och_rows: list[list[object]] = []
    oau_rows: list[list[object]] = []
    osc_rows: list[list[object]] = []

    def scan_sheet(workbook_path: Path, first_data_row: int) -> None:
        log.info("Scanning %s from row %s ...", workbook_path.name, first_data_row)
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            ws = wb[SHEET1]
            scanned = 0
            for row in ws.iter_rows(
                min_row=first_data_row,
                min_col=1,
                max_col=RAW_COLS,
                values_only=True,
            ):
                values = row_as_master_text(first_n(row, RAW_COLS))
                scanned += 1
                if is_och_performance_row(values):
                    och_rows.append(values)
                if is_oau_row(values):
                    oau_rows.append(values)
                if is_osc_row(values):
                    osc_rows.append(first_n(values, 7))
            log.info("  Scanned %s rows", scanned)
        finally:
            wb.close()

    for idx, path in enumerate(performance_files):
        first_data_row = FIRST_DATA_ROW if idx == 0 else CONTINUATION_FIRST_DATA_ROW
        scan_sheet(path, first_data_row)

    log.info("OCH Performance filtered rows: %s", len(och_rows))
    log.info("OAU filtered rows: %s", len(oau_rows))
    log.info("OSC filtered rows: %s", len(osc_rows))
    return och_rows, oau_rows, osc_rows


def paste_rows(sheet: xw.Sheet, start_cell: str, rows: list[list[object]], width: int = RAW_COLS) -> None:
    if not rows:
        return
    start = sheet.range(start_cell)
    for offset in range(0, len(rows), PASTE_CHUNK_ROWS):
        chunk = rows[offset : offset + PASTE_CHUNK_ROWS]
        target = sheet.range((start.row + offset, start.column)).resize(len(chunk), width)
        target.number_format = "@"
        target.value = chunk


def _paste_master_performance_sheets(
    wb: xw.Book,
    och_rows: list[list[object]],
    oau_rows: list[list[object]],
) -> None:
    ws_och = wb.sheets[OCH_PERFORMANCE_SHEET]
    och_used_last = max(ws_och.used_range.last_cell.row, OCH_PASTE_START_ROW)
    ws_och.range(f"C{OCH_PASTE_START_ROW}:J{och_used_last}").clear_contents()
    paste_rows(ws_och, f"C{OCH_PASTE_START_ROW}", och_rows)
    log.info("  Pasted %s rows -> %s C:J", len(och_rows), OCH_PERFORMANCE_SHEET)

    ws_oau = wb.sheets[OAU_SHEET]
    oau_used_last = max(ws_oau.used_range.last_cell.row, OAU_PASTE_START_ROW)
    ws_oau.range(f"A{OAU_PASTE_START_ROW}:H{oau_used_last}").clear_contents()
    paste_rows(ws_oau, f"A{OAU_PASTE_START_ROW}", oau_rows)
    log.info("  Pasted %s rows -> %s A:H", len(oau_rows), OAU_SHEET)


def _refresh_master_omsp_bu_on_workbook(wb: xw.Book, omsp_path: Path) -> None:
    omsp_ref = external_ref(omsp_path, OMSP_TEMPLATE_SHEET)
    ws = wb.sheets[MASTER_OMSP_BU_SHEET]
    last_row = max(ws.used_range.last_cell.row, MASTER_OMSP_BU_START_ROW)
    fill_formula_column(
        ws,
        MASTER_OMSP_BU_COL,
        last_row,
        formula_master_bu(MASTER_OMSP_BU_START_ROW, omsp_ref),
        start_row=MASTER_OMSP_BU_START_ROW,
    )
    log.info(
        "  Master %s!%s refreshed rows %s:%s",
        MASTER_OMSP_BU_SHEET,
        MASTER_OMSP_BU_COL,
        MASTER_OMSP_BU_START_ROW,
        last_row,
    )


def apply_master_performance_updates(
    master_path: Path,
    och_rows: list[list[object]],
    oau_rows: list[list[object]],
    omsp_path: Path | None = None,
    *,
    paste_performance: bool = True,
) -> None:
    """Open master once: paste performance data, optional BU refresh, recalc, save."""
    log.info("Updating master performance sheets: %s", master_path)
    with excel_session() as app:
        wb = app.books.open(str(master_path), update_links=False)
        try:
            if paste_performance:
                _paste_master_performance_sheets(wb, och_rows, oau_rows)
            if omsp_path is not None:
                if not omsp_path.is_file():
                    raise FileNotFoundError(f"Missing OMSP/DWDM workbook for BU refresh: {omsp_path}")
                log.info("Refreshing master %s column %s -> %s", MASTER_OMSP_BU_SHEET, MASTER_OMSP_BU_COL, omsp_path.name)
                _refresh_master_omsp_bu_on_workbook(wb, omsp_path)
            log.info("Recalculating master formulas ...")
            force_excel_calculate(app, wb)
            wb.save()
            log.info("Master saved.")
        finally:
            wb.close()


def update_master_performance(master_path: Path, och_rows: list[list[object]], oau_rows: list[list[object]]) -> None:
    apply_master_performance_updates(master_path, och_rows, oau_rows, omsp_path=None)


def formula_master_bu(row: int, omsp_ref: str) -> str:
    return (
        f"=IFERROR(INDEX({omsp_ref}$AB:$AB,MATCH({MASTER_OMSP_BU_MATCH_COL}{row},{omsp_ref}$D:$D,0)),"
        f"INDEX({omsp_ref}$AC:$AC,MATCH({MASTER_OMSP_BU_MATCH_COL}{row},{omsp_ref}$C:$C,0)))"
    )


def refresh_master_omsp_bu_column(master_path: Path, omsp_path: Path) -> None:
    apply_master_performance_updates(
        master_path,
        [],
        [],
        omsp_path=omsp_path,
        paste_performance=False,
    )


def optional_omsp_path(cfg: dict, config_path: Path) -> Path | None:
    return optional_pipeline_path(cfg, config_path, "dwdm_omsp_output")


def update_omsp_osc_performance(omsp_path: Path | None, osc_rows: list[list[object]]) -> None:
    if omsp_path is None:
        log.info("Skipping OMSP OSC paste; pipeline.dwdm_omsp_output is not configured.")
        return
    if not omsp_path.is_file():
        raise FileNotFoundError(f"Missing OMSP/DWDM workbook: {omsp_path}")

    log.info("Updating OMSP OSC sheet: %s", omsp_path)
    with excel_session() as app:
        wb = app.books.open(str(omsp_path), update_links=False)
        try:
            ws = wb.sheets[OSC_SHEET]
            used_last = max(ws.used_range.last_cell.row, OSC_PASTE_START_ROW)
            ws.range(f"C{OSC_PASTE_START_ROW}:I{used_last}").clear_contents()
            if used_last > OSC_PASTE_START_ROW:
                ws.range(f"J{OSC_PASTE_START_ROW + 1}:Q{used_last}").clear_contents()
            paste_rows(ws, f"C{OSC_PASTE_START_ROW}", osc_rows, width=7)
            if len(osc_rows) > 1:
                last_row = OSC_PASTE_START_ROW + len(osc_rows) - 1
                formula_source = ws.range(f"J{OSC_PASTE_START_ROW}:Q{OSC_PASTE_START_ROW}")
                formula_target = ws.range(f"J{OSC_PASTE_START_ROW}:Q{last_row}")
                formula_source.api.AutoFill(formula_target.api)
            app.calculate()
            wb.save()
            log.info("  Pasted %s rows -> %s C:I", len(osc_rows), OSC_SHEET)
        finally:
            wb.close()


def run_fast_master_paste(
    config_path: Path,
    *,
    use_daily: bool = True,
    archive_daily: bool = True,
) -> int:
    cfg = load_yaml(config_path)
    performance_files = resolve_performance_files(
        cfg,
        config_path,
        use_daily=use_daily,
        archive_daily=archive_daily,
    )
    paths = build_pipeline_paths(cfg, config_path)
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    omsp_path = optional_omsp_path(cfg, config_path)

    log.info("Fast mode: filtering Current Performance files (%s workbook(s)).", len(performance_files))
    och_rows, oau_rows, osc_rows = collect_filtered_rows(performance_files)

    master_backup: Path | None = None
    if pipe.get("make_master_backup", True):
        master_backup = backup_path_for(paths.master_workbook, "_BACKUP")
        shutil.copy2(paths.master_workbook, master_backup)
        log.info("Master backup -> %s", master_backup.name)

    try:
        apply_master_performance_updates(
            paths.master_workbook,
            och_rows,
            oau_rows,
            omsp_path=omsp_path,
        )
        update_omsp_osc_performance(omsp_path, osc_rows)
    except Exception:
        if master_backup is not None and restore_workbook_from_backup(master_backup, paths.master_workbook):
            log.error("Restored master workbook from backup after failure.")
        raise
    return 0


def verify_output(output_path: Path, performance_files: list[Path]) -> None:
    if not performance_files:
        raise ValueError("No Current Performance workbooks were provided")
    workbooks = [load_workbook(path, read_only=True, data_only=False) for path in performance_files]
    out_wb = load_workbook(output_path, read_only=True, data_only=False)
    try:
        worksheets = [wb[SHEET1] for wb in workbooks]
        main_ws = worksheets[0]
        expected_rows = main_ws.max_row + sum(max(0, ws.max_row - 1) for ws in worksheets[1:])
        ws = out_wb[SHEET1]
        if out_wb.sheetnames[:3] != [SHEET1, SHEET2, SHEET3]:
            raise AssertionError(f"Unexpected sheets: {out_wb.sheetnames}")

        actual_rows = ws.max_row
        if actual_rows is None:
            log.info("Output worksheet dimension has no max_row; counting rows by streaming ...")
            actual_rows = sum(1 for _ in ws.iter_rows(values_only=True))
            out_wb.close()
            out_wb = load_workbook(output_path, read_only=True, data_only=False)
            ws = out_wb[SHEET1]

        if actual_rows != expected_rows:
            raise AssertionError(f"Expected {expected_rows} rows, got {actual_rows}")

        headers = [ws.cell(MAIN_HEADER_ROW, c).value for c in range(1, OUTPUT_COLS + 1)]
        expected_filter_headers = [FILTER_HEADERS[9], FILTER_HEADERS[10], FILTER_HEADERS[11]]
        if headers[8:11] != expected_filter_headers:
            raise AssertionError(f"Bad filter headers in I:K row 8: {headers[8:11]}")

        for row_idx in (FIRST_DATA_ROW, actual_rows):
            formula_i = ws.cell(row_idx, 9).value
            formulas_jk = [ws.cell(row_idx, c).value for c in range(10, 12)]
            has_i_formula = isinstance(formula_i, ArrayFormula) or (
                isinstance(formula_i, str) and formula_i.startswith("=")
            )
            has_jk_formulas = all(isinstance(v, str) and v.startswith("=") for v in formulas_jk)
            if not has_i_formula or not has_jk_formulas:
                raise AssertionError(
                    f"Missing formulas in I:K row {row_idx}: {[formula_i] + formulas_jk}"
                )
        if len(worksheets) > 1:
            continuation_first = [
                worksheets[1].cell(CONTINUATION_FIRST_DATA_ROW, c).value
                for c in range(1, RAW_COLS + 1)
            ]
            pasted_first = [ws.cell(main_ws.max_row + 1, c).value for c in range(1, RAW_COLS + 1)]
            if pasted_first != continuation_first:
                raise AssertionError("Continuation first data row was not pasted at the expected output row")
        total_cell = ws.cell(6, 1).value
        expected_total = total_records_text(actual_rows - (FIRST_DATA_ROW - 1))
        if total_cell != expected_total:
            raise AssertionError(f"Expected row 6 total {expected_total!r}, got {total_cell!r}")
        if len(worksheets) > 1 and re.match(r"Monitored Object", str(ws.cell(main_ws.max_row + 1, 1).value or "")):
            raise AssertionError("Continuation header row appears to have been duplicated")
        log.info("Verification OK.")
    finally:
        for wb in workbooks:
            wb.close()
        out_wb.close()


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    parser = argparse.ArgumentParser(description="Build merged Current Performance Data workbook")
    parser.add_argument("--config", type=Path, required=True, help="Path to ingest.yml")
    parser.add_argument("--template", type=Path, help="Desired-output workbook to copy styles from")
    parser.add_argument("--output", type=Path, help="Output .xlsx path")
    args = parser.parse_args(argv)

    config_path = args.config.resolve()
    if not config_path.is_file():
        log.error("Config not found: %s", config_path)
        return 1

    try:
        cfg, performance_files = performance_paths(config_path)
        paths = build_pipeline_paths(cfg, config_path)
        week_label = str(cfg["week_label"]).strip()
        template_path = (
            args.template.resolve()
            if args.template
            else optional_pipeline_path(cfg, config_path, "current_performance_template")
        )
        if template_path is None:
            log.info("No Current Performance template configured; using the main workbook for styles.")
        output_path = (
            args.output.resolve()
            if args.output
            else optional_pipeline_path(cfg, config_path, "current_performance_output")
            or output_default(config_path, week_label).resolve()
        )
        input_paths = {path.resolve() for path in performance_files}
        protected_paths = input_paths | ({template_path.resolve()} if template_path else set())
        if output_path.resolve() in protected_paths:
            log.error("Refusing to overwrite input/template workbook: %s", output_path)
            return 1
        och_rows, oau_rows, osc_rows = build_output(
            performance_files,
            template_path.resolve() if template_path else None,
            output_path,
        )
        verify_output(output_path, performance_files)
        pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
        master_backup: Path | None = None
        if pipe.get("make_master_backup", True):
            master_backup = backup_path_for(paths.master_workbook, "_BACKUP")
            shutil.copy2(paths.master_workbook, master_backup)
            log.info("Master backup -> %s", master_backup.name)
        omsp_path = optional_omsp_path(cfg, config_path)
        try:
            apply_master_performance_updates(
                paths.master_workbook,
                och_rows,
                oau_rows,
                omsp_path=omsp_path,
            )
            update_omsp_osc_performance(omsp_path, osc_rows)
        except Exception:
            if master_backup is not None and restore_workbook_from_backup(
                master_backup, paths.master_workbook
            ):
                log.error("Restored master workbook from backup after failure.")
            raise
        return 0
    except Exception:
        log.exception("build_current_performance_output failed")
        return 1


if __name__ == "__main__":
    sys.exit(main())
