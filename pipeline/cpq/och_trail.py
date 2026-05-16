"""
OCh Trail Integrity -- Formula Workbook Builder
=================================================
Uses ingest.yml (--config).

USAGE
-----
    python och_trail.py --config ingest.yml

REQUIREMENTS
------------
    pip install openpyxl pyyaml
"""

from __future__ import annotations

import importlib.util
from pathlib import Path


def _bootstrap_import_paths() -> None:
    pipeline_dir = next(p for p in Path(__file__).resolve().parents if p.name == "pipeline")
    spec = importlib.util.spec_from_file_location("_pathsetup", pipeline_dir / "_pathsetup.py")
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    module.setup(__file__)


_bootstrap_import_paths()

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pipeline_config import (
    WB_OCH_TRAIL,
    WB_OMS_COMPUTED,
    build_pipeline_paths,
    get_sheet_spec,
    idx_map_from_headers,
    load_yaml,
)

log = logging.getLogger("och_builder")

OCH_TRAILS_SHEET = "OCh Trails"
OCH_ROUTES_SHEET = "OCh Routes"
OMS_ROUTES_SHEET = "OMS Routes"

OMS_ROUTE_KEYS = ["oms_name", "span"]


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)


def header_style(cell, bg="1F4E79", fg="FFFFFF"):
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, color=fg, name="Arial", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border("FFFFFF")


def data_style(cell, bg=None):
    cell.font = Font(name="Arial", size=9)
    cell.border = _border()
    if bg:
        cell.fill = _fill(bg)


def formula_style(cell, bg=None):
    cell.font = Font(name="Arial", size=9, color="1F4E79")
    cell.border = _border()
    if bg:
        cell.fill = _fill(bg)


def _load_rows(filepath: Path, sheet_name: str, data_only: bool) -> list[tuple]:
    wb = load_workbook(str(filepath), read_only=True, data_only=data_only)
    try:
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def read_sheet(filepath: Path, sheet_name: str, spec) -> tuple[list, list]:
    rows = _load_rows(filepath, sheet_name, data_only=True)
    if not rows:
        return [], []
    hi = spec.header_row_1based - 1
    if len(rows) <= hi:
        return [], []
    headers = list(rows[hi])
    data = [list(r) for r in rows[hi + 1 :] if any(v is not None for v in r)]

    if not data and len(rows) > hi + 1:
        log.warning("  %r: re-reading with data_only=False", sheet_name)
        rows = _load_rows(filepath, sheet_name, data_only=False)
        if not rows:
            return headers, data
        hi = spec.header_row_1based - 1
        headers = list(rows[hi])
        data = [list(r) for r in rows[hi + 1 :] if any(v is not None for v in r)]
        if data:
            log.info("  %r: recovered %s rows", sheet_name, len(data))

    return headers, data


def read_oms_ref(filepath: Path, workbooks: dict) -> list[tuple]:
    spec = get_sheet_spec(workbooks, WB_OMS_COMPUTED, OMS_ROUTES_SHEET)
    rows = _load_rows(filepath, OMS_ROUTES_SHEET, data_only=True)
    if not rows:
        rows = _load_rows(filepath, OMS_ROUTES_SHEET, data_only=False)
    if not rows:
        log.warning("OMS Routes empty")
        return []

    hi = spec.header_row_1based - 1
    headers = [str(v).strip() if v is not None else "" for v in rows[hi]]
    log.info("  OMS Routes header row Excel %s", hi + 1)

    try:
        mp = idx_map_from_headers(headers, spec.columns, OMS_ROUTE_KEYS)
        name_idx = mp["oms_name"]
        span_idx = mp["span"]
    except (KeyError, ValueError) as e:
        log.error("%s", e)
        return []

    pairs = []
    for row in rows[hi + 1 :]:
        if not any(v is not None for v in row):
            continue
        name = row[name_idx] if name_idx < len(row) else None
        span = row[span_idx] if span_idx < len(row) else None
        pairs.append((name, span))

    log.info("  OMSRef pairs: %s", len(pairs))
    return pairs


def write_raw_sheet(wb, title, headers, data, header_bg="1F4E79", alt_bg="DCE6F1"):
    ws = wb.create_sheet(title=title)
    for ci, h in enumerate(headers, 1):
        header_style(ws.cell(row=1, column=ci, value=h), bg=header_bg)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, row in enumerate(data, start=2):
        bg = alt_bg if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            data_style(ws.cell(row=ri, column=ci, value=val), bg=bg)

    for ci, h in enumerate(headers, 1):
        col_vals = [str(h or "")] + [
            str(r[ci - 1] or "") if (ci - 1) < len(r) else "" for r in data
        ]
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(v) for v in col_vals) + 2, 50)

    log.info("  %s: %s rows, %s cols", title, len(data), len(headers))
    return ws


def write_och_routes_sheet(wb, headers, data):
    ws = wb.create_sheet(title=OCH_ROUTES_SHEET)
    display_headers = list(headers)[:5]
    while len(display_headers) < 5:
        display_headers.append("")
    display_headers[4] = "OMS Span"

    for ci, h in enumerate(display_headers, 1):
        bg = "BF8F00" if ci == 5 else "1F4E79"
        header_style(ws.cell(row=1, column=ci, value=h), bg=bg)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, row in enumerate(data, start=2):
        alt = ri % 2 == 0
        for ci in range(1, 5):
            val = row[ci - 1] if (ci - 1) < len(row) else None
            data_style(ws.cell(row=ri, column=ci, value=val), bg="DCE6F1" if alt else None)
        f_span = f'=IFERROR(INDEX(OMSRef!$B:$B,MATCH(D{ri},OMSRef!$A:$A,0)),"N/A")'
        formula_style(ws.cell(row=ri, column=5, value=f_span), bg="FFF2CC" if alt else "FFFEF5")

    for ci, w in enumerate([55, 20, 20, 60, 35], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    log.info("  %s: %s rows, col E = formula", OCH_ROUTES_SHEET, len(data))
    return ws


def write_oms_ref_sheet(wb, pairs):
    ws = wb.create_sheet(title="OMSRef")
    for ci, h in enumerate(["OMS Name", "Span"], 1):
        header_style(ws.cell(row=1, column=ci, value=h), bg="085041")

    for ri, (name, span) in enumerate(pairs, start=2):
        bg = "E1F5EE" if ri % 2 == 0 else None
        data_style(ws.cell(row=ri, column=1, value=name), bg=bg)
        data_style(ws.cell(row=ri, column=2, value=span), bg=bg)

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 35
    ws.sheet_state = "hidden"
    log.info("  OMSRef: %s rows (hidden)", len(pairs))
    return ws


def force_recalc(output_path: Path) -> None:
    try:
        import xlwings as xw
    except ImportError:
        log.warning("xlwings not installed — open the file and press Ctrl+S once.")
        return
    log.info("Opening in Excel to force recalculation ...")
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        wb = app.books.open(str(output_path))
        app.calculate()
        wb.save()
        wb.close()
        log.info("   Recalculation done ✓")
    except Exception as e:
        log.warning("   Recalc failed: %s", e)
    finally:
        app.screen_updating = True
        app.quit()


def run(config_path: Path) -> int:
    cfg = load_yaml(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    workbooks = cfg.get("workbooks")
    if not isinstance(workbooks, dict):
        log.error("workbooks mapping required")
        return 1

    trails_spec = get_sheet_spec(workbooks, WB_OCH_TRAIL, OCH_TRAILS_SHEET)
    routes_spec = get_sheet_spec(workbooks, WB_OCH_TRAIL, OCH_ROUTES_SHEET)

    och_path = paths.och_trail_source
    oms_path = paths.oms_computed_out
    output_path = paths.och_computed_out

    if not och_path.exists():
        log.error("OCh file not found: %s", och_path)
        return 1
    if not oms_path.exists():
        log.error("OMS computed not found: %s", oms_path)
        return 1

    trails_headers, trails_data = read_sheet(och_path, OCH_TRAILS_SHEET, trails_spec)
    routes_headers, routes_data = read_sheet(och_path, OCH_ROUTES_SHEET, routes_spec)

    if not trails_data and not routes_data:
        log.error("Both OCh sheets empty — aborting.")
        return 1

    oms_pairs = read_oms_ref(oms_path, workbooks)
    if not oms_pairs:
        log.error("No OMSRef data — aborting.")
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    write_raw_sheet(wb, OCH_TRAILS_SHEET, trails_headers, trails_data)
    write_och_routes_sheet(wb, routes_headers, routes_data)
    write_oms_ref_sheet(wb, oms_pairs)
    wb.remove(default)

    wb.save(str(output_path))
    log.info("Saved -> %s", output_path.resolve())
    force_recalc(output_path)
    log.info("Done → %s", output_path.resolve())
    return 0


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    p = argparse.ArgumentParser(description="Build OCh computed workbook")
    p.add_argument("--config", type=Path, required=True)
    args = p.parse_args(argv)
    cfg_path = args.config.resolve()
    if not cfg_path.is_file():
        log.error("Config not found: %s", cfg_path)
        return 1
    return run(cfg_path)


if __name__ == "__main__":
    sys.exit(main())
