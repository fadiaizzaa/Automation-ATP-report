"""
OMS Trail Integrity -- Formula Workbook Builder
================================================
Uses ingest.yml (--config) for paths, header rows, and column titles.

USAGE
-----
    python oms_trail.py --config ingest.yml

REQUIREMENTS
------------
    pip install openpyxl pyyaml
"""

from __future__ import annotations

import importlib.util
from pathlib import Path


def _bootstrap_import_paths() -> None:
    pipeline_dir = next(p for p in Path(__file__).resolve().parents if p.name == "pipeline")
    spec = importlib.util.spec_from_file_location("_pathsetup", pipeline_dir / "_pathsetup.py")
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    module.setup(__file__)


_bootstrap_import_paths()

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pipeline_config import (
    WB_FIBER_COMPUTED,
    WB_OMS_TRAIL,
    build_pipeline_paths,
    get_sheet_spec,
    idx_map_from_headers,
    load_yaml,
)

log = logging.getLogger("oms_builder")

OMS_TRAILS_SHEET = "OMS Trails"
OMS_ROUTES_SHEET = "OMS Routes"
EXCLUDE_ROUTE_KEYWORDS = ("Obverse", "Reverse")

FIBER_NAME_KEYS = ["name_star", "finally_dwdm_span"]


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)


def header_style(cell, bg="1F4E79", fg="FFFFFF"):
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, color=fg, name="Arial", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border("FFFFFF")


def data_style(cell, bg=None):
    cell.font = Font(name="Arial", size=9)
    cell.border = _border()
    if bg:
        cell.fill = _fill(bg)


def formula_style(cell, bg=None):
    cell.font = Font(name="Arial", size=9, color="1F4E79")
    cell.border = _border()
    if bg:
        cell.fill = _fill(bg)


def _read_sheet_rows(filepath: Path, sheet_name: str, data_only: bool) -> list[tuple]:
    wb = load_workbook(str(filepath), read_only=True, data_only=data_only)
    try:
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def read_sheet(filepath: Path, sheet_name: str, spec) -> tuple[list, list]:
    rows = _read_sheet_rows(filepath, sheet_name, data_only=True)
    if not rows:
        return [], []
    hi = spec.header_row_1based - 1
    if len(rows) <= hi:
        return [], []
    headers = list(rows[hi])
    data = [list(r) for r in rows[hi + 1 :] if any(v is not None for v in r)]

    if not data and len(rows) > hi + 1:
        log.warning(
            "  Sheet %r: 0 data rows with cached values; re-reading data_only=False",
            sheet_name,
        )
        rows = _read_sheet_rows(filepath, sheet_name, data_only=False)
        if not rows:
            return headers, data
        hi = spec.header_row_1based - 1
        headers = list(rows[hi])
        data = [list(r) for r in rows[hi + 1 :] if any(v is not None for v in r)]
        if data:
            log.info("  Sheet %r: recovered %s rows", sheet_name, len(data))

    return headers, data


def filter_routes(data: list[list], col_idx: int, exclude_keywords: tuple[str, ...]):
    keywords_lower = [k.lower() for k in exclude_keywords]
    kept, removed = [], 0
    for row in data:
        if col_idx < len(row) and row[col_idx] is not None:
            val = str(row[col_idx]).lower()
        else:
            val = ""
        if any(kw in val for kw in keywords_lower):
            removed += 1
        else:
            kept.append(row)
    return kept, removed


def read_fiber_ref(filepath: Path, workbooks: dict) -> list[tuple]:
    spec = get_sheet_spec(workbooks, WB_FIBER_COMPUTED, "Fiber")
    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    try:
        ws = wb["Fiber"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        log.warning("Fiber sheet is empty!")
        return []

    hi = spec.header_row_1based - 1
    headers = [str(v).strip() if v is not None else "" for v in rows[hi]]
    log.info("  Fiber header Excel row %s, %s columns", hi + 1, len(headers))

    try:
        mp = idx_map_from_headers(headers, spec.columns, FIBER_NAME_KEYS)
        d_idx = mp["name_star"]
        av_idx = mp["finally_dwdm_span"]
    except (KeyError, ValueError) as e:
        log.error("%s", e)
        return []

    pairs = []
    for row in rows[hi + 1 :]:
        if not any(v is not None for v in row):
            continue
        d = row[d_idx] if d_idx < len(row) else None
        av = row[av_idx] if av_idx < len(row) else None
        pairs.append((d, av))

    log.info("  FiberRef pairs: %s rows", len(pairs))
    return pairs


def write_raw_sheet(wb, title, headers, data, header_bg="1F4E79", alt_bg="DCE6F1"):
    ws = wb.create_sheet(title=title)
    for ci, h in enumerate(headers, 1):
        header_style(ws.cell(row=1, column=ci, value=h), bg=header_bg)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, row in enumerate(data, start=2):
        bg = alt_bg if ri % 2 == 0 else None
        for ci, val in enumerate(row, 1):
            data_style(ws.cell(row=ri, column=ci, value=val), bg=bg)

    for ci, h in enumerate(headers, 1):
        col_vals = [str(h or "")] + [
            str(r[ci - 1] or "") if (ci - 1) < len(r) else "" for r in data
        ]
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(v) for v in col_vals) + 2, 50)

    log.info("  %s: %s rows, %s cols", title, len(data), len(headers))
    return ws


def write_oms_routes_sheet(wb, headers, data):
    ws = wb.create_sheet(title=OMS_ROUTES_SHEET)
    display_headers = list(headers)[:5]
    while len(display_headers) < 5:
        display_headers.append("")
    display_headers[4] = "Span"

    for ci, h in enumerate(display_headers, 1):
        bg = "BF8F00" if ci == 5 else "1F4E79"
        header_style(ws.cell(row=1, column=ci, value=h), bg=bg)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for ri, row in enumerate(data, start=2):
        alt = ri % 2 == 0
        for ci in range(1, 5):
            val = row[ci - 1] if (ci - 1) < len(row) else None
            data_style(ws.cell(row=ri, column=ci, value=val), bg="DCE6F1" if alt else None)
        formula = f'=IFERROR(INDEX(FiberRef!$B:$B,MATCH(D{ri},FiberRef!$A:$A,0)),"N/A")'
        formula_style(ws.cell(row=ri, column=5, value=formula), bg="FFF2CC" if alt else "FFFEF5")

    for ci, w in enumerate([60, 20, 20, 40, 35], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    log.info("  %s: %s rows (filtered), col E = formula", OMS_ROUTES_SHEET, len(data))
    return ws


def write_fiber_ref_sheet(wb, pairs):
    ws = wb.create_sheet(title="FiberRef")
    for ci, h in enumerate(["Name* (D)", "FINALDWDMSpanValue (AV)"], 1):
        header_style(ws.cell(row=1, column=ci, value=h), bg="085041")

    for ri, (d, av) in enumerate(pairs, start=2):
        bg = "E1F5EE" if ri % 2 == 0 else None
        data_style(ws.cell(row=ri, column=1, value=d), bg=bg)
        data_style(ws.cell(row=ri, column=2, value=av), bg=bg)

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 35
    ws.sheet_state = "hidden"
    log.info("  FiberRef: %s rows (hidden)", len(pairs))
    return ws


def force_recalc(output_path: Path) -> None:
    try:
        import xlwings as xw
    except ImportError:
        log.warning("xlwings not installed — open the file and press Ctrl+S once.")
        return
    log.info("Opening in Excel to force recalculation ...")
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        wb = app.books.open(str(output_path))
        app.calculate()
        wb.save()
        wb.close()
        log.info("   Recalculation done ✓")
    except Exception as e:
        log.warning("   Recalc failed: %s — open the file and press Ctrl+S once.", e)
    finally:
        app.screen_updating = True
        app.quit()


def run(config_path: Path) -> int:
    cfg = load_yaml(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    workbooks = cfg.get("workbooks")
    if not isinstance(workbooks, dict):
        log.error("workbooks mapping required")
        return 1

    trails_spec = get_sheet_spec(workbooks, WB_OMS_TRAIL, OMS_TRAILS_SHEET)
    routes_spec = get_sheet_spec(workbooks, WB_OMS_TRAIL, OMS_ROUTES_SHEET)

    oms_path = paths.oms_trail_source
    fiber_path = paths.fiber_computed_out
    output_path = paths.oms_computed_out

    if not oms_path.exists():
        log.error("OMS file not found: %s", oms_path)
        return 1
    if not fiber_path.exists():
        log.error("Fiber computed not found: %s", fiber_path)
        return 1

    log.info("OMS source  : %s", oms_path.name)
    log.info("Fiber source: %s", fiber_path.name)

    trails_headers, trails_data = read_sheet(oms_path, OMS_TRAILS_SHEET, trails_spec)
    routes_headers, routes_data = read_sheet(oms_path, OMS_ROUTES_SHEET, routes_spec)

    try:
        hdr_route = routes_headers
        rt_idx = idx_map_from_headers(hdr_route, routes_spec.columns, ["route_type"])["route_type"]
    except (KeyError, ValueError) as e:
        log.error("%s", e)
        return 1

    routes_data, removed = filter_routes(routes_data, rt_idx, EXCLUDE_ROUTE_KEYWORDS)
    log.info(
        "  Removed (Obverse/Reverse): %s  |  Kept: %s",
        removed,
        len(routes_data),
    )

    if not trails_data and not routes_data:
        log.error("Both OMS sheets empty — aborting.")
        return 1

    fiber_pairs = read_fiber_ref(fiber_path, workbooks)
    if not fiber_pairs:
        log.error("No FiberRef pairs — aborting.")
        return 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    blank_default = wb.active
    write_raw_sheet(wb, OMS_TRAILS_SHEET, trails_headers, trails_data)
    write_oms_routes_sheet(wb, routes_headers, routes_data)
    write_fiber_ref_sheet(wb, fiber_pairs)
    wb.remove(blank_default)

    wb.save(str(output_path))
    log.info("Saved -> %s", output_path.resolve())
    force_recalc(output_path)
    log.info("Done → %s", output_path.resolve())
    return 0


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    p = argparse.ArgumentParser(description="Build OMS computed workbook")
    p.add_argument("--config", type=Path, required=True)
    args = p.parse_args(argv)
    cfg_path = args.config.resolve()
    if not cfg_path.is_file():
        log.error("Config not found: %s", cfg_path)
        return 1
    return run(cfg_path)


if __name__ == "__main__":
    sys.exit(main())
