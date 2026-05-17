"""
Master Workbook Updater  (xlwings version)
==========================================
Pastes computed data from OCh/OMS computed workbooks into the master
OTS ATP NMS workbook using xlwings — which drives real Excel, so ALL
Excel features are preserved.

Column positions for OCh Routes and OMS Routes come from ingest.yml
(`och_computed` / `oms_computed`). OMS Count still pastes the first five
physical columns (A–E) of OMS Routes; OMS Trails pastes the first 24
columns when present.

USAGE
-----
  python pasting_cpq2.py --config ingest.yml

REQUIREMENTS
------------
  pip install xlwings openpyxl pyyaml
  Microsoft Excel must be installed (Windows or Mac).
"""

from __future__ import annotations

import importlib.util
from pathlib import Path


def _bootstrap_import_paths() -> None:
    pipeline_dir = next(p for p in Path(__file__).resolve().parents if p.name == "pipeline")
    spec = importlib.util.spec_from_file_location("_pathsetup", pipeline_dir / "_pathsetup.py")
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    module.setup(__file__)


_bootstrap_import_paths()

import argparse
import logging
import shutil
import sys
from pathlib import Path

import xlwings as xw # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.utils import get_column_letter, column_index_from_string # type: ignore

from excel_refs import external_ref, fill_formula_column, force_excel_calculate, restore_workbook_from_backup
from pipeline_config import (
    WB_OCH_COMPUTED,
    WB_OMS_COMPUTED,
    build_pipeline_paths,
    get_sheet_spec,
    idx_map_from_headers,
    load_yaml,
)

log = logging.getLogger("master_updater")

OCH_ROUTES_SHEET = "OCh Routes"
OCH_COUNT_SHEET = "OCh Count"
FIBER_OCH_SHEET = "OCh"
OMS_ROUTES_SHEET = "OMS Routes"
OMS_TRAILS_SHEET = "OMS Trails"

OCH_ROUTE_KEYS = ["route_col_1", "route_col_2", "route_col_3", "route_col_4", "oms_span"]
OMS_ROUTE_KEYS = ["oms_name", "trail_name", "span"]
LIB_SHEET = "Lib"
MASTER_OMS_T_SHEET = "OMS_T"


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def read_sheet(filepath: Path, sheet_name: str, header_row_idx: int):
    """Read sheet values. header_row_idx is 0-based. Returns (headers, data_rows)."""
    wb = load_workbook(str(filepath), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows or len(rows) <= header_row_idx:
        return [], []
    headers = list(rows[header_row_idx])
    data = [list(r) for r in rows[header_row_idx + 1 :] if any(v is not None for v in r)]
    return headers, data


def clear_cols(sheet: xw.Sheet, col_letters: list, start_row: int, nrows: int):
    for col in col_letters:
        sheet.range(f"{col}{start_row}:{col}{start_row + nrows - 1}").clear_contents()


def paste_single_col(sheet: xw.Sheet, data: list, src_idx: int, dest_col: str, start_row: int):
    values = [[row[src_idx] if src_idx < len(row) else None] for row in data]
    if values:
        end_row = start_row + len(values) - 1
        sheet.range(f"{dest_col}{start_row}:{dest_col}{end_row}").value = values


def paste_block(sheet: xw.Sheet, data: list, src_start: int, src_end: int, dest_col: str, start_row: int):
    ncols = src_end - src_start + 1
    values = [
        [row[i] if i < len(row) else None for i in range(src_start, src_end + 1)]
        for row in data
    ]
    if not values:
        return
    dest_end_col = get_column_letter(column_index_from_string(dest_col) + ncols - 1)
    end_row = start_row + len(values) - 1
    sheet.range(f"{dest_col}{start_row}:{dest_end_col}{end_row}").value = values


def to_flat(vals):
    """Flatten xlwings range value into a plain list."""
    if vals is None:
        return []
    if isinstance(vals, list):
        return [v[0] if isinstance(v, list) else v for v in vals]
    return [vals]


def to_rows(vals):
    """Normalize xlwings range value into a list of rows."""
    if vals is None:
        return []
    if not isinstance(vals, list):
        return [[vals]]
    if vals and not isinstance(vals[0], list):
        return [vals]
    return vals


def parse_ne_name(ne_name: str):
    """
    Parse "{NE_ID}_{prefix}_{SiteName}" into (int ne_id, str site).
    Returns (None, None) if unparseable.
    """
    if not ne_name or not isinstance(ne_name, str):
        return None, None
    parts = ne_name.split("_", 2)
    if len(parts) < 3:
        return None, None
    try:
        ne_id = int(parts[0])
    except ValueError:
        return None, None
    return ne_id, parts[2]


def last_row_in_col(sheet: xw.Sheet, col: str, min_row: int = 1) -> int:
    row = sheet.range(f"{col}{sheet.cells.last_cell.row}").end("up").row
    return max(row, min_row)


def update_lib_from_oms_t(wb: xw.Book) -> int:
    """
    Update Lib columns N/O/P using NE names from refreshed OMS_T columns L and R.
    Returns number of new NE rows added.
    """
    ws_omst = wb.sheets[MASTER_OMS_T_SHEET]
    ws_lib = wb.sheets[LIB_SHEET]

    log.info("Updating Lib N/O/P from OMS_T L/R ...")
    last_l = last_row_in_col(ws_omst, "L", min_row=4)
    last_r = last_row_in_col(ws_omst, "R", min_row=4)
    last_row = max(last_l, last_r)
    if last_row < 4:
        log.info("  OMS_T has no data rows for Lib update.")
        return 0

    log.info("  OMS_T data: rows 4 to %s", last_row)
    src = to_flat(ws_omst.range(f"L4:L{last_row}").value)
    snk = to_flat(ws_omst.range(f"R4:R{last_row}").value)

    all_ne_names = {
        name.strip()
        for name in src + snk
        if name and isinstance(name, str) and "_" in name
    }
    log.info("  Unique NE names found: %s", len(all_ne_names))

    parsed = {}
    unparseable = []
    for name in all_ne_names:
        ne_id, site = parse_ne_name(name)
        if ne_id is not None:
            parsed.setdefault(ne_id, (name, site))
        else:
            unparseable.append(name)

    if unparseable:
        log.warning("  Could not parse %s name(s): %s", len(unparseable), unparseable[:5])

    last_nop_row = last_row_in_col(ws_lib, "N", min_row=1)
    existing_vals = ws_lib.range(f"N2:O{last_nop_row}").value if last_nop_row >= 2 else None
    existing_nop = set()
    for row in to_rows(existing_vals):
        ne_name_val, ne_id_val = (row if isinstance(row, list) else [row, None])
        if ne_id_val is not None and str(ne_name_val) != "NEName":
            try:
                existing_nop.add(int(ne_id_val))
            except (ValueError, TypeError):
                pass

    missing_nop = {ne_id: v for ne_id, v in parsed.items() if ne_id not in existing_nop}
    log.info("  Existing Lib N/O/P IDs: %s, missing: %s", len(existing_nop), len(missing_nop))

    if not missing_nop:
        log.info("  Lib N/O/P already up to date.")
        return 0

    rows_to_write = [
        [ne_name, ne_id, site]
        for ne_id, (ne_name, site) in sorted(missing_nop.items())
    ]
    start_row = last_nop_row + 1
    end_row = start_row + len(rows_to_write) - 1
    ws_lib.range(f"N{start_row}:P{end_row}").value = rows_to_write
    log.info("  Added %s NE(s) to Lib N/O/P rows %s:%s", len(rows_to_write), start_row, end_row)
    return len(rows_to_write)


def formula_och_count_h(row: int, och_ref: str) -> str:
    return f"=INDEX({och_ref}$H:$H,MATCH(A{row},{och_ref}$D:$D,0))"


def refresh_och_count_column_h(
    wb: xw.Book,
    fiber_computed_path: Path,
    data_start: int,
    och_row_count: int,
) -> None:
    if och_row_count <= 0:
        return
    if not fiber_computed_path.is_file():
        log.warning("Fiber computed workbook missing; skipping OCh Count column H refresh.")
        return

    och_ref = external_ref(fiber_computed_path, FIBER_OCH_SHEET)
    ws = wb.sheets[OCH_COUNT_SHEET]
    last_row = max(ws.used_range.last_cell.row, data_start + och_row_count - 1)
    fill_formula_column(ws, "H", last_row, formula_och_count_h(data_start, och_ref), start_row=data_start)
    log.info("  OCh Count column H refreshed -> %s rows %s:%s", FIBER_OCH_SHEET, data_start, last_row)


def force_recalc(output_path: Path):
    log.info("   Opening master for recalculation …")
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = None
    try:
        wb = app.books.open(str(output_path), update_links=False)
        force_excel_calculate(app, wb)
        wb.save()
        log.info("   Recalculation done ✓")
    except Exception as e:
        log.warning("   Recalc failed: %s — open the file and press Ctrl+S once.", e)
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        try:
            app.screen_updating = True
            app.quit()
        except Exception:
            pass


def _log_column_mapping(sheet_label: str, keys: list[str], idx_map: dict[str, int]) -> None:
    for k in keys:
        if k in idx_map:
            log.info("  %s  %s -> column %s", sheet_label, k, get_column_letter(idx_map[k] + 1))


def _validate_och_route_contiguous(idx_map: dict[str, int]) -> bool:
    r1 = idx_map["route_col_1"]
    r2 = idx_map["route_col_2"]
    r3 = idx_map["route_col_3"]
    r4 = idx_map["route_col_4"]
    if r2 != r1 + 1 or r3 != r2 + 1 or r4 != r3 + 1:
        log.error(
            "OCh Routes: route_col_1..4 must be consecutive columns; got %s, %s, %s, %s",
            r1,
            r2,
            r3,
            r4,
        )
        return False
    return True


def _validate_oms_order(idx_map: dict[str, int]) -> bool:
    om = idx_map["oms_name"]
    tr = idx_map["trail_name"]
    sp = idx_map["span"]
    if not (tr > om and sp > om):
        log.error(
            "OMS Routes: trail_name and span must be to the right of oms_name; got oms=%s trail=%s span=%s",
            om,
            tr,
            sp,
        )
        return False
    return True


def _validate_oms_trail_span_adjacent(idx_map: dict[str, int]) -> bool:
    tr = idx_map["trail_name"]
    sp = idx_map["span"]
    if sp != tr + 1:
        log.error(
            "OMS Routes: trail_name and span must be adjacent columns for master N:O; got trail=%s span=%s",
            tr,
            sp,
        )
        return False
    return True


def run(config_path: Path) -> int:
    cfg = load_yaml(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    workbooks = cfg.get("workbooks")
    if not isinstance(workbooks, dict):
        log.error("workbooks mapping required")
        return 1

    och_spec = get_sheet_spec(workbooks, WB_OCH_COMPUTED, OCH_ROUTES_SHEET)
    oms_routes_spec = get_sheet_spec(workbooks, WB_OMS_COMPUTED, OMS_ROUTES_SHEET)
    oms_trails_spec = get_sheet_spec(workbooks, WB_OMS_COMPUTED, OMS_TRAILS_SHEET)

    master_path = paths.master_workbook
    och_path = paths.och_computed_out
    oms_path = paths.oms_computed_out
    log.info("Master workbook path: %s", master_path)

    fiber_path = paths.fiber_computed_out
    for p in (master_path, och_path, oms_path, fiber_path):
        if not p.exists():
            log.error("File not found: %s", p)
            return 1

    och_h = och_spec.header_row_1based - 1
    oms_r_h = oms_routes_spec.header_row_1based - 1
    oms_t_h = oms_trails_spec.header_row_1based - 1

    log.info("Reading computed files …")
    och_headers, och_routes = read_sheet(och_path, OCH_ROUTES_SHEET, header_row_idx=och_h)
    log.info("  OCh Routes : %s rows", len(och_routes))

    oms_headers, oms_routes = read_sheet(oms_path, OMS_ROUTES_SHEET, header_row_idx=oms_r_h)
    log.info("  OMS Routes : %s rows", len(oms_routes))

    oms_th_headers, oms_trails = read_sheet(oms_path, OMS_TRAILS_SHEET, header_row_idx=oms_t_h)
    log.info("  OMS Trails : %s rows", len(oms_trails))

    if not any([och_routes, oms_routes, oms_trails]):
        log.error("All source sheets are empty — aborting.")
        return 1

    try:
        och_idx = idx_map_from_headers(och_headers, och_spec.columns, OCH_ROUTE_KEYS)
    except (KeyError, ValueError) as e:
        log.error("OCh Routes column map: %s", e)
        return 1

    if not _validate_och_route_contiguous(och_idx):
        return 1

    try:
        om_idx = idx_map_from_headers(oms_headers, oms_routes_spec.columns, OMS_ROUTE_KEYS)
    except (KeyError, ValueError) as e:
        log.error("OMS Routes column map: %s", e)
        return 1

    if not _validate_oms_order(om_idx):
        return 1
    if not _validate_oms_trail_span_adjacent(om_idx):
        return 1

    _log_column_mapping("OCh Routes", OCH_ROUTE_KEYS, och_idx)
    _log_column_mapping("OMS Routes", OMS_ROUTE_KEYS, om_idx)

    r1 = och_idx["route_col_1"]
    r4 = och_idx["route_col_4"]
    span_i = och_idx["oms_span"]
    om_i = om_idx["oms_name"]
    tr_i = om_idx["trail_name"]
    sp_i = om_idx["span"]

    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    make_backup = pipe.get("make_master_backup", True)
    backup_path: Path | None = None
    if make_backup:
        backup_path = master_path.with_stem(master_path.stem + "_BACKUP")
        shutil.copy2(master_path, backup_path)
        log.info("Backup → %s", backup_path.name)

    log.info("Opening master in Excel: %s …", master_path.name)
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False

    DATA_START = 2
    OMS_T_START = 4

    try:
        wb = app.books.open(str(master_path), update_links=False)

        log.info("Pasting → OCh Count …")
        sht = wb.sheets[OCH_COUNT_SHEET]
        och_n = 0

        if och_routes:
            och_n = len(och_routes)
            clear_cols(sht, ["A", "B", "C", "D", "G"], DATA_START, och_n)
            paste_block(sht, och_routes, r1, r4, "A", DATA_START)
            paste_single_col(sht, och_routes, span_i, "G", DATA_START)
            log.info("  OCh Routes → A–D, G : %s rows", och_n)

        refresh_och_count_column_h(wb, paths.fiber_computed_out, DATA_START, och_n)

        if oms_routes:
            n = len(oms_routes)
            clear_cols(sht, ["M", "N", "O"], DATA_START, n)
            paste_single_col(sht, oms_routes, om_i, "M", DATA_START)
            paste_block(sht, oms_routes, tr_i, sp_i, "N", DATA_START)
            log.info("  OMS Routes → M, N–O : %s rows", n)

        log.info("Pasting → OMS Count …")
        sht = wb.sheets["OMS Count"]
        if oms_routes:
            n = len(oms_routes)
            clear_cols(sht, ["B", "C", "D", "E", "F"], DATA_START, n)
            n_oms_cols = min(5, len(oms_headers))
            if len(oms_headers) < 5:
                log.warning(
                    "OMS Routes has %s header columns; OMS Count paste uses first %s (expects 5 / A–E).",
                    len(oms_headers),
                    n_oms_cols,
                )
            if n_oms_cols > 0:
                paste_block(sht, oms_routes, 0, n_oms_cols - 1, "B", DATA_START)
            log.info("  OMS Routes → B–F : %s rows", n)

        log.info("Pasting → OMS_T …")
        sht = wb.sheets["OMS_T"]
        if oms_trails:
            n = len(oms_trails)
            n_trail_src = min(24, len(oms_th_headers))
            if len(oms_th_headers) < 24:
                log.warning(
                    "OMS Trails has %s columns (expected >= 24 for C–Z paste)",
                    len(oms_th_headers),
                )
            clear_cols(sht, [get_column_letter(c) for c in range(3, 27)], OMS_T_START, n)
            if n_trail_src > 0:
                paste_block(sht, oms_trails, 0, n_trail_src - 1, "C", OMS_T_START)
            log.info("  OMS Trails → C–Z : %s rows, %s cols", n, n_trail_src)

        added_lib_rows = update_lib_from_oms_t(wb)
        log.info("Lib update complete: %s new row(s).", added_lib_rows)

        log.info("Saving …")
        wb.save()
        wb.close()

        force_recalc(master_path)

        log.info("Master updated → %s", master_path.name)
        log.info("Sheets updated : OCh Count, OMS Count, OMS_T, Lib")
        log.info("Done → %s", master_path.resolve())
        return 0
    except Exception as e:
        if backup_path is not None and restore_workbook_from_backup(backup_path, master_path):
            log.error("Restored master workbook from backup after failure.")
        log.error("Error: %s", e)
        raise
    finally:
        try:
            app.screen_updating = True
            app.quit()
        except Exception as ex:
            log.warning("Excel app cleanup (first instance): %s", ex)


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    p = argparse.ArgumentParser(description="Paste computed CPQ data into master workbook")
    p.add_argument("--config", type=Path, required=True)
    args = p.parse_args(argv)
    cfg_path = args.config.resolve()
    if not cfg_path.is_file():
        log.error("Config not found: %s", cfg_path)
        return 1
    try:
        return run(cfg_path)
    except Exception:
        log.exception("pasting_cpq2 failed")
        return 1


if __name__ == "__main__":
    sys.exit(main())
