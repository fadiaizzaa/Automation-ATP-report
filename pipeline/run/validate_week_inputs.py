#!/usr/bin/env python3
"""Validate prepared weekly input workbook layouts before computed steps run."""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pipeline_config import (
    WB_NETWORK_RESOURCE,
    WB_OCH_TRAIL,
    WB_OMS_TRAIL,
    WB_SERVICE_ROUTING,
    build_pipeline_paths,
    get_sheet_spec,
    idx_map_from_headers,
    load_yaml,
)

log = logging.getLogger("validate_week_inputs")

NRR_CURRENT_KEYS = [
    "row_no", "phase", "resource_status", "name_star", "source_site", "sink_site",
    "type_star", "zero_dispersion_area", "domain_star", "direction", "distance_km",
    "attenuation_db", "attenuation_coefficient", "dispersion_ps_nm",
    "dispersion_coefficient", "pmd", "dgd", "margin_db", "other_loss_db",
    "user_cost", "remarks",
]
NRR_PREVIOUS_KEYS = ["remarks", "prev_span_value", "prev_remark", "prev_span_count"]

SRR_OCH_KEYS = [
    "och_c", "och_e", "och_g", "och_h", "och_m", "och_o", "och_p", "och_q",
    "och_s", "och_ac", "och_ae", "och_af", "och_ag", "och_ai",
]
SRR_E2E_KEYS = [
    "e2e_c", "e2e_d", "e2e_e", "e2e_h", "e2e_j", "e2e_m", "e2e_o",
    "e2e_s", "e2e_u", "e2e_x", "e2e_z", "e2e_ad", "e2e_ae", "e2e_ag",
    "e2e_aj", "e2e_al", "e2e_ao", "e2e_as", "e2e_av", "e2e_ax", "e2e_ba",
]


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def _headers(path: Path, sheet_name: str, header_row_1based: int) -> list[Any]:
    if not path.is_file():
        raise FileNotFoundError(str(path))
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{path.name}: missing sheet {sheet_name!r}")
        ws = wb[sheet_name]
        if ws.max_row < header_row_1based:
            raise ValueError(
                f"{path.name}/{sheet_name}: too short for header row {header_row_1based}"
            )
        return [cell.value for cell in ws[header_row_1based]]
    finally:
        wb.close()


def _validate_sheet(
    workbook_label: str,
    path: Path,
    sheet_name: str,
    header_row_1based: int,
    columns: dict[str, str],
    required_keys: list[str],
) -> bool:
    try:
        headers = _headers(path, sheet_name, header_row_1based)
        idx_map_from_headers(headers, columns, required_keys)
    except (OSError, KeyError, ValueError) as e:
        log.error(
            "%s / %s failed layout validation (header row %s, file %s): %s",
            workbook_label,
            sheet_name,
            header_row_1based,
            path,
            e,
        )
        return False
    log.info(
        "%s / %s OK (%s required columns)",
        workbook_label,
        sheet_name,
        len(required_keys),
    )
    return True


def _validate_sheet_exists(
    workbook_label: str,
    path: Path,
    sheet_name: str,
    header_row_1based: int,
) -> bool:
    try:
        _headers(path, sheet_name, header_row_1based)
    except (OSError, ValueError) as e:
        log.error(
            "%s / %s failed sheet validation (header row %s, file %s): %s",
            workbook_label,
            sheet_name,
            header_row_1based,
            path,
            e,
        )
        return False
    log.info("%s / %s OK", workbook_label, sheet_name)
    return True


def run(config_path: Path) -> int:
    cfg = load_yaml(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    workbooks = cfg.get("workbooks")
    if not isinstance(workbooks, dict):
        log.error("Config must define workbooks:")
        return 1

    ok = True

    nrr_spec = get_sheet_spec(workbooks, WB_NETWORK_RESOURCE, "Fiber")
    ok &= _validate_sheet(
        "Network Resource Statistics (current)",
        paths.current_network_resource,
        "Fiber",
        nrr_spec.header_row_1based,
        nrr_spec.columns,
        NRR_CURRENT_KEYS,
    )
    ok &= _validate_sheet(
        "Network Resource Statistics (previous)",
        paths.previous_network_resource,
        "Fiber",
        nrr_spec.header_row_1based,
        nrr_spec.columns,
        NRR_PREVIOUS_KEYS,
    )

    oms_trails = get_sheet_spec(workbooks, WB_OMS_TRAIL, "OMS Trails")
    oms_routes = get_sheet_spec(workbooks, WB_OMS_TRAIL, "OMS Routes")
    ok &= _validate_sheet_exists(
        "OMS Trail Integrity",
        paths.oms_trail_source,
        "OMS Trails",
        oms_trails.header_row_1based,
    )
    ok &= _validate_sheet_exists(
        "OMS Trail Integrity",
        paths.oms_trail_source,
        "OMS Routes",
        oms_routes.header_row_1based,
    )

    och_trails = get_sheet_spec(workbooks, WB_OCH_TRAIL, "OCh Trails")
    och_routes = get_sheet_spec(workbooks, WB_OCH_TRAIL, "OCh Routes")
    ok &= _validate_sheet_exists(
        "OCh Trail Integrity",
        paths.och_trail_source,
        "OCh Trails",
        och_trails.header_row_1based,
    )
    ok &= _validate_sheet_exists(
        "OCh Trail Integrity",
        paths.och_trail_source,
        "OCh Routes",
        och_routes.header_row_1based,
    )

    srr_och = get_sheet_spec(workbooks, WB_SERVICE_ROUTING, "OCh Route")
    srr_e2e = get_sheet_spec(workbooks, WB_SERVICE_ROUTING, "Trail E2E Route")
    ok &= _validate_sheet(
        "Service Routing Report",
        paths.service_routing_raw,
        "OCh Route",
        srr_och.header_row_1based,
        srr_och.columns,
        SRR_OCH_KEYS,
    )
    ok &= _validate_sheet(
        "Service Routing Report",
        paths.service_routing_raw,
        "Trail E2E Route",
        srr_e2e.header_row_1based,
        srr_e2e.columns,
        SRR_E2E_KEYS,
    )

    if not ok:
        log.error("Input layout validation failed. Stop before computed/master updates.")
        return 1
    log.info("All prepared input layouts passed validation.")
    return 0


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    p = argparse.ArgumentParser(description="Validate prepared weekly input layouts")
    p.add_argument("--config", type=Path, required=True, help="Path to ingest YAML")
    args = p.parse_args(argv)
    cfg_path = args.config.resolve()
    if not cfg_path.is_file():
        log.error("Config not found: %s", cfg_path)
        return 1
    try:
        return run(cfg_path)
    except Exception:
        log.exception("validate_week_inputs failed")
        return 1


if __name__ == "__main__":
    sys.exit(main())
