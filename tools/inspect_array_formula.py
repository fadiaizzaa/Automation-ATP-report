from pathlib import Path
from openpyxl import load_workbook


TEMPLATE = Path(r"C:\Users\fadia\Downloads\Week 19_Current_Performance_Data_2026-05-07_16-34-38.xlsx")


def main() -> None:
    wb = load_workbook(TEMPLATE, read_only=False, data_only=False)
    ws = wb["Sheet1"]
    for coord in ["I9", "I10", "J9", "K9"]:
        value = ws[coord].value
        print(coord, type(value), repr(value))
        if hasattr(value, "__dict__"):
            print(value.__dict__)
        print("data_type", ws[coord].data_type)
    print("array_formulae", getattr(ws, "array_formulae", None))
    wb.close()


if __name__ == "__main__":
    main()
