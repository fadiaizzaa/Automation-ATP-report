from pathlib import Path
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell


TEMPLATE = Path(r"C:\Users\fadia\Downloads\Week 19_Current_Performance_Data_2026-05-07_16-34-38.xlsx")
RAW = Path(r"C:\Users\fadia\Downloads\Huawei-Fadia Izza Nabila\ATP NMS\input\Week 19\NMS\Week 19_Current Performance Data.xlsx")
RAW_1 = Path(r"C:\Users\fadia\Downloads\Huawei-Fadia Izza Nabila\ATP NMS\input\Week 19\NMS\Week 19_Current Performance Data_1.xlsx")
OUTPUT = Path(r"C:\Users\fadia\Downloads\Huawei-Fadia Izza Nabila\final-project-huawei-fadia\outputs\week19_current_performance_data_output.xlsx")

FILTER_LINE_PATTERNS = [
    "*LOG*",
    "*N30*",
    "*N40*",
    "*N50*",
    "*N60*",
    "*NS4*",
    "*NS3*",
    "*ND2*",
    "*LSX*",
    "*LDX*",
    "*ELOM*",
    "*LSC*",
    "*LTX*",
    "*LQM*",
    "*LDC*",
]
FILTER_AMP_PATTERNS = [
    "*VA*",
    "*OAU*",
    "*OBU*",
    "*RAU*",
    "*RPC*",
    "*DAP*",
    "*MD40*",
    "*RAPXF*",
    "*MR4*",
    "*AFS*",
    "*OPU*",
    "*WSMD*",
]
FILTER_OSC_PATTERNS = ["*ST*", "*SC*"]


def excel_array(patterns: list[str]) -> str:
    return "{" + ",".join(f'"{pattern}"' for pattern in patterns) + "}"


def formula(row_idx: int, patterns: list[str]) -> str:
    return f'=SUM(COUNTIF(A{row_idx}, {excel_array(patterns)})) > 0'


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def styled_row(ws, row_values, style_cells=None):
    cells = []
    for col_idx, value in enumerate(row_values, start=1):
        cell = WriteOnlyCell(ws, value=value)
        if style_cells and col_idx <= len(style_cells):
            copy_cell_style(style_cells[col_idx - 1], cell)
        cells.append(cell)
    return cells


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    template_wb = load_workbook(TEMPLATE, read_only=False, data_only=False)
    template_ws = template_wb["Sheet1"]
    header_styles = {
        row_idx: [template_ws.cell(row_idx, col_idx) for col_idx in range(1, 12)]
        for row_idx in range(1, 9)
    }
    data_styles = [template_ws.cell(9, col_idx) for col_idx in range(1, 12)]
    widths = {
        col: template_ws.column_dimensions[col].width
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
        if template_ws.column_dimensions[col].width
    }
    template_wb.close()

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet("Sheet1")
    for col, width in widths.items():
        out_ws.column_dimensions[col].width = width

    raw_wb = load_workbook(RAW, read_only=True, data_only=False)
    raw_ws = raw_wb["Sheet1"]

    for row_idx, row in enumerate(raw_ws.iter_rows(values_only=True), start=1):
        values = list(row[:8])
        if row_idx <= 7:
            values = values + [None, None, None]
        elif row_idx == 8:
            values = values + ["Filter Line Board", "Filter Amp Board", "Filter OSC Board"]
        else:
            values = values + [
                formula(row_idx, FILTER_LINE_PATTERNS),
                formula(row_idx, FILTER_AMP_PATTERNS),
                formula(row_idx, FILTER_OSC_PATTERNS),
            ]

        styles = header_styles.get(row_idx, data_styles)
        out_ws.append(styled_row(out_ws, values, styles))

    raw_wb.close()

    next_output_row = raw_ws.max_row + 1
    raw_1_wb = load_workbook(RAW_1, read_only=True, data_only=False)
    raw_1_ws = raw_1_wb["Sheet1"]

    for source_row_idx, row in enumerate(raw_1_ws.iter_rows(values_only=True), start=1):
        if source_row_idx == 1:
            continue
        values = list(row[:8]) + [
            formula(next_output_row, FILTER_LINE_PATTERNS),
            formula(next_output_row, FILTER_AMP_PATTERNS),
            formula(next_output_row, FILTER_OSC_PATTERNS),
        ]
        out_ws.append(styled_row(out_ws, values, data_styles))
        next_output_row += 1

    raw_1_wb.close()

    out_wb.create_sheet("Sheet2")
    out_wb.create_sheet("Sheet3")
    out_wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
