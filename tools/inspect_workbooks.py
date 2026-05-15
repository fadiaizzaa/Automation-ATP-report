from pathlib import Path
from openpyxl import load_workbook


FILES = [
    Path(r"C:\Users\fadia\Downloads\Week 19_Current_Performance_Data_2026-05-07_16-34-38.xlsx"),
    Path(r"C:\Users\fadia\Downloads\Huawei-Fadia Izza Nabila\ATP NMS\input\Week 19\NMS\Week 19_Current Performance Data.xlsx"),
    Path(r"C:\Users\fadia\Downloads\Huawei-Fadia Izza Nabila\ATP NMS\input\Week 19\NMS\Week 19_Current Performance Data_1.xlsx"),
]


def main() -> None:
    for file_path in FILES:
        print(f"FILE\t{file_path}")
        wb = load_workbook(file_path, read_only=True, data_only=False)
        for ws in wb.worksheets:
            formula_cells = []
            non_empty_formula_cols = set()
            sample_headers = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 5), values_only=True):
                sample_headers.append(row[: min(ws.max_column, 12)])
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cells.append(cell.coordinate)
                        non_empty_formula_cols.add(cell.column)
                        if len(formula_cells) >= 20:
                            break
                if len(formula_cells) >= 20:
                    break
            print(
                f"SHEET\t{ws.title}\trows={ws.max_row}\tcols={ws.max_column}\t"
                f"formula_sample={formula_cells}\tformula_cols={sorted(non_empty_formula_cols)}"
            )
            print(f"HEADERS\t{sample_headers}")
        wb.close()


if __name__ == "__main__":
    main()
