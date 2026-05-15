from pathlib import Path
from openpyxl import load_workbook


TEMPLATE = Path(r"C:\Users\fadia\Downloads\Week 19_Current_Performance_Data_2026-05-07_16-34-38.xlsx")
RAW = Path(r"C:\Users\fadia\Downloads\Huawei-Fadia Izza Nabila\ATP NMS\input\Week 19\NMS\Week 19_Current Performance Data.xlsx")
RAW_1 = Path(r"C:\Users\fadia\Downloads\Huawei-Fadia Izza Nabila\ATP NMS\input\Week 19\NMS\Week 19_Current Performance Data_1.xlsx")


def show_rows(path: Path, max_col: int = 11) -> None:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb["Sheet1"]
    print(f"FILE\t{path}")
    first_rows = []
    tail_rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = list(row[:max_col])
        if row_idx <= 12:
            first_rows.append((row_idx, values))
        tail_rows.append((row_idx, values))
        if len(tail_rows) > 4:
            tail_rows.pop(0)
    for row_idx, values in first_rows:
        print(row_idx, values)
    print("TAIL")
    for row_idx, values in tail_rows:
        print(row_idx, values)
    wb.close()


def main() -> None:
    show_rows(TEMPLATE)
    show_rows(RAW)
    show_rows(RAW_1)


if __name__ == "__main__":
    main()
