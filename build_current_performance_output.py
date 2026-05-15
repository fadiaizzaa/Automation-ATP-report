#!/usr/bin/env python3
"""
Build the merged, formula-ready Current Performance Data workbook.

This is a separate/manual helper. It does not run from run_week_pipeline.py.

Example:
    python build_current_performance_output.py --config ingest.yml --template "C:\\Users\\fadia\\Downloads\\Week 19_Current_Performance_Data_2026-05-07_16-34-38.xlsx"
"""

from __future__ import annotations

import argparse
import logging
import re
import shutil
import sys
from copy import copy
from datetime import date, datetime, time
from pathlib import Path
from typing import Iterable

import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.worksheet.formula import ArrayFormula

from pipeline_config import build_pipeline_paths, expand_config_value, load_yaml, resolve_under


log = logging.getLogger("current_performance_builder")

SHEET1 = "Sheet1"
SHEET2 = "Sheet2"
SHEET3 = "Sheet3"

RAW_COLS = 8
OUTPUT_COLS = 11
MAIN_HEADER_ROW = 8
FIRST_DATA_ROW = 9
CONTINUATION_FIRST_DATA_ROW = 2

FILTER_HEADERS = {
    9: "Filter Line Board",
    10: "Filter Amp Board",
    11: "Filter OSC Board",
}

FILTER_LINE_PATTERNS = [
    "*LOG*",
    "*N30*",
    "*N40*",
    "*N50*",
    "*N60*",
    "*NS4*",
    "*NS3*",
    "*ND2*",
    "*LSX*",
    "*LDX*",
    "*ELOM*",
    "*LSC*",
    "*LTX*",
    "*LQM*",
    "*LDC*",
]
FILTER_AMP_PATTERNS = [
    "*VA*",
    "*OAU*",
    "*OBU*",
    "*RAU*",
    "*RPC*",
    "*DAP*",
    "*MD40*",
    "*RAPXF*",
    "*MR4*",
    "*AFS*",
    "*OPU*",
    "*WSMD*",
]
FILTER_OSC_PATTERNS = ["*ST*", "*SC*"]

OCH_PERFORMANCE_SHEET = "OCH Performance"
OAU_SHEET = "OAU"
OCH_PASTE_START_ROW = 7
OAU_PASTE_START_ROW = 6
PASTE_CHUNK_ROWS = 20_000

OCH_PERFORMANCE_EVENTS = {
    "LSIOPCUR",
    "LSOOPCUR",
    "FEC_BEF_COR_ER",
    "FEC_AFT_COR_ER",
    "FEC_BEF_CORER_FLOAT",
    "FEC_AFT_CORER_FLOAT",
    "TDCCUR",
    "DGDCUR",
}
OAU_PERFORMANCE_EVENTS = {"SUMIOPCUR", "SUMOOPCUR"}


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def excel_array(patterns: list[str]) -> str:
    return "{" + ",".join(f'"{pattern}"' for pattern in patterns) + "}"


def filter_formula(row_idx: int, patterns: list[str]) -> str:
    return f'=SUM(COUNTIF(A{row_idx}, {excel_array(patterns)})) > 0'


def filter_formulas(row_idx: int) -> list[str]:
    return [
        ArrayFormula(f"I{row_idx}", filter_formula(row_idx, FILTER_LINE_PATTERNS)),
        filter_formula(row_idx, FILTER_AMP_PATTERNS),
        filter_formula(row_idx, FILTER_OSC_PATTERNS),
    ]


def event_code(value: object) -> str:
    text = str(value or "").strip().upper()
    return text.split("(", 1)[0].strip()


def pattern_match(value: object, patterns: list[str]) -> bool:
    text = str(value or "").upper()
    return any(pattern.strip("*").upper() in text for pattern in patterns)


def is_och_performance_row(values: list[object]) -> bool:
    return pattern_match(values[0] if values else None, FILTER_LINE_PATTERNS) and event_code(
        values[1] if len(values) > 1 else None
    ) in OCH_PERFORMANCE_EVENTS


def is_oau_row(values: list[object]) -> bool:
    return pattern_match(values[0] if values else None, FILTER_AMP_PATTERNS) and event_code(
        values[1] if len(values) > 1 else None
    ) in OAU_PERFORMANCE_EVENTS


def copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def styled_row(ws, row_values: list[object], style_cells=None):
    cells = []
    for col_idx, value in enumerate(row_values, start=1):
        cell = WriteOnlyCell(ws, value=value)
        if style_cells and col_idx <= len(style_cells):
            copy_cell_style(style_cells[col_idx - 1], cell)
        cells.append(cell)
    return cells


def first_n(values: Iterable[object], n: int) -> list[object]:
    out = list(values)[:n]
    if len(out) < n:
        out.extend([None] * (n - len(out)))
    return out


def as_master_text(value: object) -> object:
    """Keep performance pasted values as text so master formulas behave like manual paste."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    return str(value)


def row_as_master_text(row: list[object]) -> list[object]:
    return [as_master_text(v) for v in first_n(row, RAW_COLS)]


def total_records_text(record_count: int) -> str:
    return f"Total {record_count:,} Records"


def output_default(config_path: Path, week_label: str) -> Path:
    return config_path.parent / "outputs" / f"{week_label}_Current_Performance_Data_output.xlsx"


def performance_paths(config_path: Path) -> tuple[dict, Path, Path]:
    cfg = load_yaml(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    main = paths.nms_week_dir / f"{paths.week_label}_Current Performance Data.xlsx"
    continuation = paths.nms_week_dir / f"{paths.week_label}_Current Performance Data_1.xlsx"
    return cfg, main, continuation


def optional_pipeline_path(cfg: dict, config_path: Path, key: str) -> Path | None:
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    raw = pipe.get(key)
    if not isinstance(raw, str) or not raw.strip():
        return None
    return resolve_under(config_path.parent, expand_config_value(raw.strip(), cfg))


def backup_path_for(path: Path, suffix: str) -> Path:
    backup = path.with_stem(path.stem + suffix)
    if not backup.exists():
        return backup
    for i in range(2, 1000):
        candidate = path.with_stem(f"{path.stem}{suffix}_{i}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Could not find available backup filename for {path}")


def read_template_styles(template_path: Path):
    wb = load_workbook(template_path, read_only=False, data_only=False)
    try:
        if SHEET1 not in wb.sheetnames:
            raise ValueError(f"Template workbook is missing {SHEET1!r}")
        ws = wb[SHEET1]
        header_styles = {
            row_idx: [ws.cell(row_idx, col_idx) for col_idx in range(1, OUTPUT_COLS + 1)]
            for row_idx in range(1, MAIN_HEADER_ROW + 1)
        }
        data_styles = [ws.cell(FIRST_DATA_ROW, col_idx) for col_idx in range(1, OUTPUT_COLS + 1)]
        widths = {
            letter: ws.column_dimensions[letter].width
            for letter in "ABCDEFGHIJK"
            if ws.column_dimensions[letter].width
        }
        return header_styles, data_styles, widths
    finally:
        wb.close()


def build_output(
    main_path: Path,
    continuation_path: Path,
    template_path: Path,
    output_path: Path,
) -> tuple[list[list[object]], list[list[object]]]:
    for path, label in [
        (main_path, "main Current Performance workbook"),
        (continuation_path, "continuation Current Performance workbook"),
        (template_path, "template workbook"),
    ]:
        if not path.is_file():
            raise FileNotFoundError(f"Missing {label}: {path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    header_styles, data_styles, widths = read_template_styles(template_path)

    main_wb = load_workbook(main_path, read_only=True, data_only=False)
    continuation_wb = load_workbook(continuation_path, read_only=True, data_only=False)
    try:
        main_ws = main_wb[SHEET1]
        continuation_ws = continuation_wb[SHEET1]
        och_rows: list[list[object]] = []
        oau_rows: list[list[object]] = []

        final_last_row = main_ws.max_row + max(0, continuation_ws.max_row - 1)
        final_records = final_last_row - (FIRST_DATA_ROW - 1)
        log.info("Main rows: %s", main_ws.max_row)
        log.info("Continuation rows: %s", continuation_ws.max_row)
        log.info("Final rows: %s", final_last_row)
        log.info("Final records: %s", final_records)

        out_wb = Workbook(write_only=True)
        out_ws = out_wb.create_sheet(SHEET1)
        for col, width in widths.items():
            out_ws.column_dimensions[col].width = width

        for row_idx, row in enumerate(main_ws.iter_rows(values_only=True), start=1):
            values = first_n(row, RAW_COLS)
            if row_idx == 6:
                values[0] = total_records_text(final_records)
            if row_idx < MAIN_HEADER_ROW:
                values.extend([None, None, None])
            elif row_idx == MAIN_HEADER_ROW:
                values.extend([FILTER_HEADERS[9], FILTER_HEADERS[10], FILTER_HEADERS[11]])
            else:
                values.extend(filter_formulas(row_idx))
                raw_values = row_as_master_text(first_n(row, RAW_COLS))
                if is_och_performance_row(raw_values):
                    och_rows.append(raw_values)
                if is_oau_row(raw_values):
                    oau_rows.append(raw_values)

            styles = header_styles.get(row_idx, data_styles)
            out_ws.append(styled_row(out_ws, values, styles))

        next_output_row = main_ws.max_row + 1
        for source_row_idx, row in enumerate(continuation_ws.iter_rows(values_only=True), start=1):
            if source_row_idx < CONTINUATION_FIRST_DATA_ROW:
                continue
            values = first_n(row, RAW_COLS)
            text_values = row_as_master_text(values)
            if is_och_performance_row(text_values):
                och_rows.append(text_values)
            if is_oau_row(text_values):
                oau_rows.append(text_values)
            values.extend(filter_formulas(next_output_row))
            out_ws.append(styled_row(out_ws, values, data_styles))
            next_output_row += 1

        out_wb.create_sheet(SHEET2)
        out_wb.create_sheet(SHEET3)
        out_wb.save(output_path)
        log.info("Saved -> %s", output_path)
        log.info("OCH Performance filtered rows: %s", len(och_rows))
        log.info("OAU filtered rows: %s", len(oau_rows))
        return och_rows, oau_rows
    finally:
        main_wb.close()
        continuation_wb.close()


def collect_filtered_rows(main_path: Path, continuation_path: Path) -> tuple[list[list[object]], list[list[object]]]:
    """Fast path: read raw A:H rows and collect only rows needed by the master."""
    for path, label in [
        (main_path, "main Current Performance workbook"),
        (continuation_path, "continuation Current Performance workbook"),
    ]:
        if not path.is_file():
            raise FileNotFoundError(f"Missing {label}: {path}")

    och_rows: list[list[object]] = []
    oau_rows: list[list[object]] = []

    def scan_sheet(workbook_path: Path, first_data_row: int) -> None:
        log.info("Scanning %s from row %s ...", workbook_path.name, first_data_row)
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            ws = wb[SHEET1]
            scanned = 0
            for row in ws.iter_rows(
                min_row=first_data_row,
                min_col=1,
                max_col=RAW_COLS,
                values_only=True,
            ):
                values = row_as_master_text(first_n(row, RAW_COLS))
                scanned += 1
                if is_och_performance_row(values):
                    och_rows.append(values)
                if is_oau_row(values):
                    oau_rows.append(values)
            log.info("  Scanned %s rows", scanned)
        finally:
            wb.close()

    scan_sheet(main_path, FIRST_DATA_ROW)
    scan_sheet(continuation_path, CONTINUATION_FIRST_DATA_ROW)

    log.info("OCH Performance filtered rows: %s", len(och_rows))
    log.info("OAU filtered rows: %s", len(oau_rows))
    return och_rows, oau_rows


def paste_rows(sheet: xw.Sheet, start_cell: str, rows: list[list[object]]) -> None:
    if not rows:
        return
    start = sheet.range(start_cell)
    for offset in range(0, len(rows), PASTE_CHUNK_ROWS):
        chunk = rows[offset : offset + PASTE_CHUNK_ROWS]
        target = sheet.range((start.row + offset, start.column)).resize(len(chunk), RAW_COLS)
        target.number_format = "@"
        target.value = chunk


def update_master_performance(master_path: Path, och_rows: list[list[object]], oau_rows: list[list[object]]) -> None:
    log.info("Updating master performance sheets: %s", master_path)
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = None
    try:
        wb = app.books.open(str(master_path), update_links=False)

        ws_och = wb.sheets[OCH_PERFORMANCE_SHEET]
        och_used_last = max(ws_och.used_range.last_cell.row, OCH_PASTE_START_ROW)
        ws_och.range(f"C{OCH_PASTE_START_ROW}:J{och_used_last}").clear_contents()
        paste_rows(ws_och, f"C{OCH_PASTE_START_ROW}", och_rows)
        log.info("  Pasted %s rows -> %s C:J", len(och_rows), OCH_PERFORMANCE_SHEET)

        ws_oau = wb.sheets[OAU_SHEET]
        oau_used_last = max(ws_oau.used_range.last_cell.row, OAU_PASTE_START_ROW)
        ws_oau.range(f"A{OAU_PASTE_START_ROW}:H{oau_used_last}").clear_contents()
        paste_rows(ws_oau, f"A{OAU_PASTE_START_ROW}", oau_rows)
        log.info("  Pasted %s rows -> %s A:H", len(oau_rows), OAU_SHEET)

        log.info("Recalculating master formulas ...")
        app.calculate()
        wb.save()
        log.info("Master saved.")
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        app.screen_updating = True
        app.quit()


def run_fast_master_paste(config_path: Path) -> int:
    cfg, main_path, continuation_path = performance_paths(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}

    log.info("Fast mode: filtering raw Current Performance files directly.")
    och_rows, oau_rows = collect_filtered_rows(main_path, continuation_path)

    if pipe.get("make_master_backup", True):
        backup = backup_path_for(paths.master_workbook, "_PERFORMANCE_BACKUP")
        shutil.copy2(paths.master_workbook, backup)
        log.info("Master backup -> %s", backup.name)

    update_master_performance(paths.master_workbook, och_rows, oau_rows)
    return 0


def verify_output(output_path: Path, main_path: Path, continuation_path: Path) -> None:
    main_wb = load_workbook(main_path, read_only=True, data_only=False)
    continuation_wb = load_workbook(continuation_path, read_only=True, data_only=False)
    out_wb = load_workbook(output_path, read_only=True, data_only=False)
    try:
        expected_rows = main_wb[SHEET1].max_row + continuation_wb[SHEET1].max_row - 1
        ws = out_wb[SHEET1]
        if out_wb.sheetnames[:3] != [SHEET1, SHEET2, SHEET3]:
            raise AssertionError(f"Unexpected sheets: {out_wb.sheetnames}")

        actual_rows = ws.max_row
        if actual_rows is None:
            log.info("Output worksheet dimension has no max_row; counting rows by streaming ...")
            actual_rows = sum(1 for _ in ws.iter_rows(values_only=True))
            out_wb.close()
            out_wb = load_workbook(output_path, read_only=True, data_only=False)
            ws = out_wb[SHEET1]

        if actual_rows != expected_rows:
            raise AssertionError(f"Expected {expected_rows} rows, got {actual_rows}")

        headers = [ws.cell(MAIN_HEADER_ROW, c).value for c in range(1, OUTPUT_COLS + 1)]
        expected_filter_headers = [FILTER_HEADERS[9], FILTER_HEADERS[10], FILTER_HEADERS[11]]
        if headers[8:11] != expected_filter_headers:
            raise AssertionError(f"Bad filter headers in I:K row 8: {headers[8:11]}")

        for row_idx in (FIRST_DATA_ROW, actual_rows):
            formula_i = ws.cell(row_idx, 9).value
            formulas_jk = [ws.cell(row_idx, c).value for c in range(10, 12)]
            has_i_formula = isinstance(formula_i, ArrayFormula) or (
                isinstance(formula_i, str) and formula_i.startswith("=")
            )
            has_jk_formulas = all(isinstance(v, str) and v.startswith("=") for v in formulas_jk)
            if not has_i_formula or not has_jk_formulas:
                raise AssertionError(
                    f"Missing formulas in I:K row {row_idx}: {[formula_i] + formulas_jk}"
                )
        continuation_first = [
            continuation_wb[SHEET1].cell(CONTINUATION_FIRST_DATA_ROW, c).value
            for c in range(1, RAW_COLS + 1)
        ]
        pasted_first = [ws.cell(main_wb[SHEET1].max_row + 1, c).value for c in range(1, RAW_COLS + 1)]
        if pasted_first != continuation_first:
            raise AssertionError("Continuation first data row was not pasted at the expected output row")
        total_cell = ws.cell(6, 1).value
        expected_total = total_records_text(actual_rows - (FIRST_DATA_ROW - 1))
        if total_cell != expected_total:
            raise AssertionError(f"Expected row 6 total {expected_total!r}, got {total_cell!r}")
        if re.match(r"Monitored Object", str(ws.cell(main_wb[SHEET1].max_row + 1, 1).value or "")):
            raise AssertionError("Continuation header row appears to have been duplicated")
        log.info("Verification OK.")
    finally:
        main_wb.close()
        continuation_wb.close()
        out_wb.close()


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    parser = argparse.ArgumentParser(description="Build merged Current Performance Data workbook")
    parser.add_argument("--config", type=Path, required=True, help="Path to ingest.yml")
    parser.add_argument("--template", type=Path, help="Desired-output workbook to copy styles from")
    parser.add_argument("--output", type=Path, help="Output .xlsx path")
    args = parser.parse_args(argv)

    config_path = args.config.resolve()
    if not config_path.is_file():
        log.error("Config not found: %s", config_path)
        return 1

    try:
        cfg, main_path, continuation_path = performance_paths(config_path)
        paths = build_pipeline_paths(cfg, config_path)
        week_label = str(cfg["week_label"]).strip()
        template_path = (
            args.template.resolve()
            if args.template
            else optional_pipeline_path(cfg, config_path, "current_performance_template")
        )
        if template_path is None:
            log.error("Set --template or pipeline.current_performance_template in ingest.yml")
            return 1
        output_path = (
            args.output.resolve()
            if args.output
            else optional_pipeline_path(cfg, config_path, "current_performance_output")
            or output_default(config_path, week_label).resolve()
        )
        if output_path in {main_path.resolve(), continuation_path.resolve(), template_path.resolve()}:
            log.error("Refusing to overwrite input/template workbook: %s", output_path)
            return 1
        och_rows, oau_rows = build_output(main_path, continuation_path, template_path.resolve(), output_path)
        verify_output(output_path, main_path, continuation_path)
        pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
        if pipe.get("make_master_backup", True):
            backup = backup_path_for(paths.master_workbook, "_BACKUP")
            shutil.copy2(paths.master_workbook, backup)
            log.info("Master backup -> %s", backup.name)
        update_master_performance(paths.master_workbook, och_rows, oau_rows)
        return 0
    except Exception:
        log.exception("build_current_performance_output failed")
        return 1


if __name__ == "__main__":
    sys.exit(main())
