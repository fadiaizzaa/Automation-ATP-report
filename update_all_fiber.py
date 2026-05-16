#!/usr/bin/env python3
"""Refresh the All Fiber workbook from prepared Fiber Table Data files."""

from __future__ import annotations

import argparse
import logging
import sys
from copy import copy
from pathlib import Path
from typing import Iterable

import xlwings as xw
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

from pipeline_config import configured_output_base, expand_config_value, load_yaml, prepared_input_root, resolve_under


log = logging.getLogger("update_all_fiber")

REPORT_SHEET_PREFIX = "report fiber table data"
LIBRARY_SHEET = "Library FIU and OSC Connection"
OVERLAY_SHEET = "Overlay Names"
OMSP_TEMPLATE_SHEET = "TemplateForSystem (2)"
HEADER_ROW = 4
DATA_ROW = 5
LIBRARY_DATA_ROW = 2
OVERLAY_DATA_ROW = 3
RAW_COLS = 24

SOURCE_PORT_TOKENS = ("FIU", "APXF", "SC", "ST")
OVERLAY_SOURCE_TOKENS = ("FIU", "APXF", "SC")
WDM_NAME_TOKENS = ("LINK", "WORKING", "PROTECTION")


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def normalize(value: object) -> str:
    return str(value or "").strip()


def contains_any(value: object, tokens: Iterable[str]) -> bool:
    text = normalize(value).upper()
    return any(token in text for token in tokens)


def find_header(headers: list[object], name: str) -> int:
    want = name.strip().lower()
    for idx, value in enumerate(headers):
        if normalize(value).lower() == want:
            return idx
    raise ValueError(f"Header {name!r} not found in Fiber Table Data headers")


def fiber_table_sort_key(path: Path) -> tuple[int, str]:
    stem = path.stem
    if stem.endswith("_1") and stem[:-2]:
        return (1, stem.lower())
    suffix = stem.rsplit("_", 1)[-1]
    if suffix.isdigit():
        return (int(suffix), stem.lower())
    return (0, stem.lower())


def fiber_table_files(nms_dir: Path, week_label: str) -> list[Path]:
    files = sorted(nms_dir.glob(f"{week_label}_Fiber Table Data*.xlsx"), key=fiber_table_sort_key)
    return [p for p in files if p.is_file() and not p.name.startswith("~$")]


def load_fiber_table_rows(files: list[Path]) -> tuple[list[object], list[list[object]]]:
    if not files:
        raise FileNotFoundError("No prepared Fiber Table Data files found in NMS folder")

    headers: list[object] | None = None
    rows: list[list[object]] = []
    for file_path in files:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            file_headers = [ws.cell(HEADER_ROW, col).value for col in range(1, RAW_COLS + 1)]
            if headers is None:
                headers = file_headers
            for row in ws.iter_rows(min_row=DATA_ROW, max_col=RAW_COLS, values_only=True):
                values = list(row)
                if any(value is not None for value in values):
                    rows.append(values)
        finally:
            wb.close()
        log.info("Read %s rows from %s", len(rows), file_path.name)

    if headers is None:
        raise RuntimeError("Could not read Fiber Table Data headers")
    return headers, rows


def find_report_sheet(wb) -> str:
    for name in wb.sheetnames:
        if name.lower().startswith(REPORT_SHEET_PREFIX):
            return name
    raise KeyError(f"Workbook is missing a sheet starting with {REPORT_SHEET_PREFIX!r}")


def clear_range(ws, start_row: int, columns: Iterable[int], end_row: int) -> None:
    if end_row < start_row:
        return
    for row_idx in range(start_row, end_row + 1):
        for col_idx in columns:
            ws.cell(row_idx, col_idx).value = None


def copy_cell_style(source, target) -> None:
    if not source.has_style:
        return
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def write_columns(
    ws,
    start_row: int,
    columns: list[int],
    rows: list[tuple[object, ...]],
    template_row: int,
) -> None:
    end_row = max(ws.max_row, start_row + len(rows) + 200)
    clear_range(ws, start_row, columns, end_row)
    for offset, values in enumerate(rows):
        row_idx = start_row + offset
        for col_idx, value in zip(columns, values):
            cell = ws.cell(row_idx, col_idx, value)
            copy_cell_style(ws.cell(template_row, col_idx), cell)


def translate_formula(formula: str, origin: str, target: str) -> str:
    try:
        return Translator(formula, origin=origin).translate_formula(target)
    except Exception:
        return formula


def fill_formulas(ws, template_row: int, start_row: int, end_row: int, columns: Iterable[int]) -> None:
    if end_row < start_row:
        return
    for col_idx in columns:
        template = ws.cell(template_row, col_idx)
        if not isinstance(template.value, str) or not template.value.startswith("="):
            continue
        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row_idx, col_idx)
            copy_cell_style(template, cell)
            cell.value = translate_formula(template.value, template.coordinate, cell.coordinate)


def reset_auto_filter(ws, ref: str) -> None:
    ws.auto_filter = AutoFilter(ref=ref)


def update_report_sheet(ws, headers: list[object], rows: list[list[object]]) -> None:
    ws.cell(3, 1).value = f"Total {len(rows):,}"
    clear_range(ws, DATA_ROW, range(1, RAW_COLS + 1), max(ws.max_row, DATA_ROW + len(rows) + 200))
    for row_offset, values in enumerate(rows):
        for col_idx, value in enumerate(values[:RAW_COLS], start=1):
            ws.cell(DATA_ROW + row_offset, col_idx).value = value
    last_row = max(HEADER_ROW, DATA_ROW + len(rows) - 1)
    reset_auto_filter(ws, f"A{HEADER_ROW}:X{last_row}")
    log.info("Pasted %s Fiber Table rows -> %s", len(rows), ws.title)


def update_library_sheet(ws, rows: list[tuple[object, object]]) -> None:
    clear_range(ws, LIBRARY_DATA_ROW + 1, range(1, 5), max(ws.max_row, LIBRARY_DATA_ROW + len(rows) + 200))
    write_columns(ws, LIBRARY_DATA_ROW, [1, 3], rows, LIBRARY_DATA_ROW)
    last_row = max(LIBRARY_DATA_ROW, LIBRARY_DATA_ROW + len(rows) - 1)
    fill_formulas(ws, LIBRARY_DATA_ROW, LIBRARY_DATA_ROW, last_row, [2, 4])
    reset_auto_filter(ws, f"A1:D{last_row}")
    log.info("Pasted %s FIU/OSC library rows", len(rows))


def update_overlay_sheet(
    ws,
    yellow_rows: list[tuple[object, object]],
    pink_rows: list[tuple[object, object, object]],
) -> None:
    clear_range(
        ws,
        OVERLAY_DATA_ROW + 1,
        range(1, 12),
        max(ws.max_row, OVERLAY_DATA_ROW + max(len(yellow_rows), len(pink_rows)) + 200),
    )
    write_columns(ws, OVERLAY_DATA_ROW, [1, 3], yellow_rows, OVERLAY_DATA_ROW)
    yellow_last = max(OVERLAY_DATA_ROW, OVERLAY_DATA_ROW + len(yellow_rows) - 1)
    fill_formulas(ws, OVERLAY_DATA_ROW, OVERLAY_DATA_ROW, yellow_last, [2, 4])

    write_columns(ws, OVERLAY_DATA_ROW, [5, 6, 8], pink_rows, OVERLAY_DATA_ROW)
    pink_last = max(OVERLAY_DATA_ROW, OVERLAY_DATA_ROW + len(pink_rows) - 1)
    fill_formulas(ws, OVERLAY_DATA_ROW, OVERLAY_DATA_ROW, pink_last, [7, 9, 10, 11])

    last_row = max(yellow_last, pink_last)
    reset_auto_filter(ws, f"A2:K{last_row}")
    log.info("Pasted %s yellow overlay rows", len(yellow_rows))
    log.info("Pasted %s pink overlay rows", len(pink_rows))


def calculate_workbook(path: Path) -> None:
    log.info("Opening in Excel to calculate ...")
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    wb = None
    try:
        wb = app.books.open(str(path), update_links=False)
        try:
            app.api.CalculateFullRebuild()
        except Exception:
            app.calculate()
        wb.save()
        log.info("Calculation done.")
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass
        app.screen_updating = True
        app.quit()


def configured_all_fiber_path(cfg: dict, config_path: Path) -> Path:
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    raw = pipe.get("all_fiber_output")
    if isinstance(raw, str) and raw.strip():
        return resolve_under(config_path.parent, expand_config_value(raw.strip(), cfg))
    output_base = configured_output_base(cfg, config_path.parent)
    return output_base / f"{str(cfg['week_label']).strip()}_All Fiber.xlsx"


def configured_omsp_path(cfg: dict, config_path: Path) -> Path:
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    raw = pipe.get("dwdm_omsp_output")
    if isinstance(raw, str) and raw.strip():
        return resolve_under(config_path.parent, expand_config_value(raw.strip(), cfg))
    output_base = configured_output_base(cfg, config_path.parent)
    return output_base / "output" / f"{str(cfg['week_label']).strip()}_OMSP_DWDMEvaluation.xlsx"


def configured_previous_nrr_path(cfg: dict, config_path: Path) -> Path:
    pipe = cfg.get("pipeline") if isinstance(cfg.get("pipeline"), dict) else {}
    raw = pipe.get("previous_network_resource_workbook")
    if isinstance(raw, str) and raw.strip():
        return resolve_under(config_path.parent, expand_config_value(raw.strip(), cfg))
    previous_week_label = str(cfg["previous_week_label"]).strip()
    output_base = configured_output_base(cfg, config_path.parent)
    return output_base.parent / previous_week_label / f"{previous_week_label}_Network Resource Statistics.xlsm"


def external_excel_ref(path: Path, sheet_name: str, cell_or_range: str) -> str:
    folder = str(path.parent)
    if not folder.endswith("\\"):
        folder += "\\"
    return f"'{folder}[{path.name}]{sheet_name}'!{cell_or_range}"


def collect_overlay_output_rows(ws) -> list[tuple[object, object, object]]:
    rows: list[tuple[object, object, object]] = []
    for row_idx in range(OVERLAY_DATA_ROW, ws.max_row + 1):
        name = ws.cell(row_idx, 5).value
        fiu_src = ws.cell(row_idx, 10).value
        fiu_snk = ws.cell(row_idx, 11).value
        if any(value not in (None, "") for value in (name, fiu_src, fiu_snk)):
            rows.append((name, fiu_src, fiu_snk))
    return rows


def copy_overlay_to_omsp(
    all_fiber_path: Path,
    omsp_path: Path,
    previous_nrr_path: Path,
) -> None:
    if not omsp_path.is_file():
        raise FileNotFoundError(f"OMSP/DWDM workbook not found: {omsp_path}")
    if not previous_nrr_path.is_file():
        raise FileNotFoundError(f"Previous NRR workbook not found: {previous_nrr_path}")

    src_wb = load_workbook(all_fiber_path, read_only=True, data_only=True)
    try:
        overlay_rows = collect_overlay_output_rows(src_wb[OVERLAY_SHEET])
    finally:
        src_wb.close()

    wb = load_workbook(omsp_path)
    try:
        ws = wb[OMSP_TEMPLATE_SHEET]
        end_row = max(ws.max_row, 9 + len(overlay_rows) + 200)
        clear_range(ws, 9, [4, 17, 18, 20, 22, 23], end_row)

        nrr_d = external_excel_ref(previous_nrr_path, "Fiber", "$D:$D")
        nrr_av = external_excel_ref(previous_nrr_path, "Fiber", "$AV:$AV")
        nrr_ar = external_excel_ref(previous_nrr_path, "Fiber", "$AR:$AR")
        nrr_as = external_excel_ref(previous_nrr_path, "Fiber", "$AS:$AS")
        lib_b = external_excel_ref(all_fiber_path, LIBRARY_SHEET, "$B:$B")
        lib_d = external_excel_ref(all_fiber_path, LIBRARY_SHEET, "$D:$D")

        for offset, (name, fiu_src, fiu_snk) in enumerate(overlay_rows):
            row_idx = 9 + offset
            ws.cell(row_idx, 17).value = name
            ws.cell(row_idx, 18).value = fiu_src
            ws.cell(row_idx, 20).value = fiu_snk
            ws.cell(row_idx, 4).value = (
                f"=INDEX({nrr_av},MATCH(Q{row_idx},{nrr_d},0))"
                f'&";"&INDEX({nrr_ar},MATCH(Q{row_idx},{nrr_d},0))'
                f'&";"&INDEX({nrr_as},MATCH(Q{row_idx},{nrr_d},0))'
            )
            ws.cell(row_idx, 22).value = f"=INDEX({lib_d},MATCH(S{row_idx},{lib_b},0))"
            ws.cell(row_idx, 23).value = f"=INDEX({lib_b},MATCH(U{row_idx},{lib_d},0))"

        last_row = max(9, 9 + len(overlay_rows) - 1)
        fill_formulas(ws, 9, 9, last_row, [3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 19, 21, 24, 25])
        ws.auto_filter.ref = f"A8:{get_column_letter(ws.max_column)}{last_row}"
        wb.save(omsp_path)
    finally:
        wb.close()
    log.info("Pasted %s Overlay Names rows -> %s", len(overlay_rows), omsp_path.name)


def run(config_path: Path) -> int:
    cfg = load_yaml(config_path)
    week_label = str(cfg["week_label"]).strip()
    output_base = configured_output_base(cfg, config_path.parent)
    nms_dir = prepared_input_root(output_base) / "NMS"
    all_fiber_path = configured_all_fiber_path(cfg, config_path)
    omsp_path = configured_omsp_path(cfg, config_path)
    previous_nrr_path = configured_previous_nrr_path(cfg, config_path)

    files = fiber_table_files(nms_dir, week_label)
    if not all_fiber_path.is_file():
        raise FileNotFoundError(f"All Fiber workbook not found: {all_fiber_path}")

    log.info("All Fiber workbook: %s", all_fiber_path)
    log.info("Fiber Table input files: %s", [p.name for p in files])
    headers, rows = load_fiber_table_rows(files)

    name_idx = find_header(headers, "Name")
    level_idx = find_header(headers, "Level/Capacity")
    source_idx = find_header(headers, "Source Port")
    sink_idx = find_header(headers, "Sink Port")

    library_rows = [
        (row[source_idx], row[sink_idx])
        for row in rows
        if contains_any(row[source_idx], SOURCE_PORT_TOKENS)
    ]
    yellow_rows = [
        (row[source_idx], row[sink_idx])
        for row in rows
        if contains_any(row[sink_idx], ("ST",)) and contains_any(row[source_idx], OVERLAY_SOURCE_TOKENS)
    ]
    pink_rows = [
        (row[name_idx], row[source_idx], row[sink_idx])
        for row in rows
        if contains_any(row[level_idx], ("WDM",)) and contains_any(row[name_idx], WDM_NAME_TOKENS)
    ]

    wb = load_workbook(all_fiber_path)
    try:
        report_ws = wb[find_report_sheet(wb)]
        update_report_sheet(report_ws, headers, rows)
        update_library_sheet(wb[LIBRARY_SHEET], library_rows)
        update_overlay_sheet(wb[OVERLAY_SHEET], yellow_rows, pink_rows)
        try:
            wb.save(all_fiber_path)
        except PermissionError as exc:
            raise PermissionError(
                f"Cannot save {all_fiber_path}. Close the workbook in Excel and run update_all_fiber.py again."
            ) from exc
    finally:
        wb.close()

    calculate_workbook(all_fiber_path)
    copy_overlay_to_omsp(all_fiber_path, omsp_path, previous_nrr_path)
    calculate_workbook(omsp_path)
    return 0


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    parser = argparse.ArgumentParser(description="Update All Fiber workbook from Fiber Table Data")
    parser.add_argument("--config", type=Path, default=Path("ingest.yml"))
    args = parser.parse_args(argv)
    try:
        return run(args.config.resolve())
    except Exception:
        log.exception("update_all_fiber failed")
        return 1


if __name__ == "__main__":
    sys.exit(main())
