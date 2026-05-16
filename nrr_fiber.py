"""
Fiber Sheet — Formula Workbook Builder
======================================
Paths and column titles come from ingest YAML (--config).

USAGE
-----
    python nrr_fiber.py --config ingest.yml

REQUIREMENTS
------------
    pip install openpyxl pyyaml
"""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from pipeline_config import (
    WB_NETWORK_RESOURCE,
    build_pipeline_paths,
    get_sheet_spec,
    idx_map_from_headers,
    load_yaml,
)

log = logging.getLogger("fiber_v2")

RAW_HEADERS = [
    "No.", "Phase*", "Resource Status", "Name*", "Source Site*", "Sink Site*",
    "Type*", "Zero-Dispersion Area", "Domain*", "Direction", "Distance(km)*",
    "Attenuation(dB)*", "Attenuation Coefficient(dB/km)*", "Dispersion(ps/nm)*",
    "Dispersion Coefficient(ps/nm/km)*", "PMD(ps/sqrt(km))*", "DGD(ps)*",
    "Margin(dB)", "Other Loss (dB)", "User Cost", "Remarks",
]

# Note: YAML 1.1 treats bare `no` as boolean — use row_no in ingest, not `no`.
NRR_COLUMN_KEYS = [
    "row_no", "phase", "resource_status", "name_star", "source_site", "sink_site",
    "type_star", "zero_dispersion_area", "domain_star", "direction", "distance_km",
    "attenuation_db", "attenuation_coefficient", "dispersion_ps_nm",
    "dispersion_coefficient", "pmd", "dgd", "margin_db", "other_loss_db",
    "user_cost", "remarks",
]

PREV_KEYS = ["remarks", "prev_span_value", "prev_remark", "prev_span_count"]

CUSTOMIZE_HEADERS = [f"Customize Tag{str(i).zfill(2)}" for i in range(1, 11)]

FORMULA_HEADERS = [
    "SpanName", "Span Count", "Remark", "Span Count Last Week", "Span Check",
    "Subnet", "Source", "Sink", "DWDM Span", "System", "Provider",
    "SysNo", "Sys", "Direction", "Prot Type", "OMS",
    "FINALDWDMSpanValue", "FINALDWDMLength",
]

EXPECTED_GENERATED_HEADERS = RAW_HEADERS + CUSTOMIZE_HEADERS + FORMULA_HEADERS

COL_D = "D"
COL_U = "U"
COL_AF = "AF"
COL_AG = "AG"
COL_AH = "AH"
COL_AI = "AI"
COL_AJ = "AJ"
COL_AK = "AK"
COL_AL = "AL"
COL_AM = "AM"
COL_AN = "AN"
COL_AO = "AO"
COL_AP = "AP"
COL_AQ = "AQ"
COL_AR = "AR"
COL_AS = "AS"
COL_AT = "AT"
COL_AU = "AU"
COL_AV = "AV"
COL_AW = "AW"

PW_U = "A"
PW_AV = "B"
PW_AH = "C"
PW_AG = "D"

SHEET_NAME = "Fiber"


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def resolve_workbook_path(raw: Path) -> Path | None:
    if raw.is_file():
        return raw
    for ext in (".xlsx", ".xlsm"):
        q = raw.with_suffix(ext)
        if q.is_file():
            return q
    return None


def load_grid(path: Path, sheet_name: str) -> list[tuple]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))
    finally:
        wb.close()


def extract_table(
    rows: list[tuple],
    header_row_1based: int,
    col_indices: dict[str, int],
    keys_order: list[str],
) -> list[list]:
    hi = header_row_1based - 1
    out: list[list] = []
    for row in rows[hi + 1 :]:
        if not row or not any(v is not None for v in row):
            continue
        line = []
        for k in keys_order:
            j = col_indices[k]
            line.append(row[j] if j < len(row) else None)
        out.append(line)
    return out


def formulas_for_row(r: int) -> dict[str, str]:
    return {
        COL_AF: f'=IFERROR(INDEX(PrevWeek!${PW_AV}:${PW_AV},MATCH({COL_U}{r},PrevWeek!${PW_U}:${PW_U},0)),"N/A")',
        COL_AG: f"=COUNTIF(${COL_AF}:${COL_AF},{COL_AF}{r})",
        COL_AH: f'=IFERROR(INDEX(PrevWeek!${PW_AH}:${PW_AH},MATCH({COL_U}{r},PrevWeek!${PW_U}:${PW_U},0)),"")',
        COL_AI: f'=IFERROR(INDEX(PrevWeek!${PW_AG}:${PW_AG},MATCH({COL_U}{r},PrevWeek!${PW_U}:${PW_U},0)),"")',
        COL_AJ: f"={COL_AG}{r}={COL_AI}{r}",
        COL_AK: f'=IFERROR(LEFT({COL_D}{r},FIND(" ",{COL_D}{r},1)-1),"")',
        COL_AL: f'=IFERROR(REPLACE(LEFT({COL_D}{r},FIND("-",{COL_D}{r},1)-1),1,FIND(" ",{COL_D}{r},1),""),"")',
        COL_AM: f'=IFERROR(REPLACE(LEFT({COL_D}{r},FIND("/",{COL_D}{r},1)-1),1,FIND("-",{COL_D}{r},1),""),"")',
        COL_AN: f'=IFERROR(IF({COL_AL}{r}<{COL_AM}{r},{COL_AL}{r}&"-"&{COL_AM}{r},{COL_AM}{r}&"-"&{COL_AL}{r}),{COL_AF}{r})',
        COL_AO: f'=IFERROR(REPLACE(LEFT({COL_D}{r},FIND("_",{COL_D}{r},1)-1),1,FIND("/",{COL_D}{r},1),""),"")',
        COL_AP: (
            f'=IFERROR(REPLACE(IF(ISNUMBER(FIND(":",{COL_D}{r})),'
            f'IF(FIND(")",{COL_D}{r})<FIND(":",{COL_D}{r}),'
            f'LEFT({COL_D}{r},FIND(")",{COL_D}{r})-1),'
            f'LEFT({COL_D}{r},FIND(":",{COL_D}{r})-1)),'
            f'LEFT({COL_D}{r},FIND(")",{COL_D}{r})-1)),'
            f'1,FIND("(",{COL_D}{r}),""),"")'
        ),
        COL_AQ: f'=IFERROR(VALUE(REPLACE({COL_AO}{r},1,FIND("-",{COL_AO}{r},1),"")),"")',
        COL_AR: (
            f'=IF(OR(ISNUMBER(SEARCH("Work",{COL_D}{r})),'
            f'ISNUMBER(SEARCH("Link",{COL_D}{r})),'
            f'ISNUMBER(SEARCH("OSP",{COL_D}{r}))),"Main","Prot")'
        ),
        COL_AS: f'=IFERROR(IF(LEFT({COL_AN}{r},FIND("-",{COL_AN}{r},1)-1)={COL_AL}{r},"A>B","B>A"),"")',
        COL_AT: f'=IF(OR(ISNUMBER(SEARCH("Work",{COL_D}{r})),ISNUMBER(SEARCH("Prot",{COL_D}{r}))),"1+1","1+0")',
        COL_AU: f'=IF(ISERROR(FIND("OMS",{COL_D}{r})),{COL_AN}{r},REPLACE({COL_D}{r},1,FIND("OMS:",{COL_D}{r})+3,""))',
        COL_AV: f'=IF({COL_AF}{r}="N/A",{COL_AN}{r},{COL_AF}{r})',
        COL_AW: f"={COL_AV}{r}&{COL_AR}{r}",
    }


def _border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)


def apply_header_style(cell, bg="1F4E79", fg="FFFFFF"):
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, color=fg, name="Arial", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border("FFFFFF")


def apply_data_style(cell, bg=None):
    cell.font = Font(name="Arial", size=9)
    cell.border = _border()
    if bg:
        cell.fill = _fill(bg)


def validate_generated_headers(ws) -> None:
    for ci, expected in enumerate(EXPECTED_GENERATED_HEADERS, 1):
        actual = ws.cell(row=2, column=ci).value
        if actual != expected:
            col = get_column_letter(ci)
            raise ValueError(
                f"Generated Fiber layout mismatch at {col}2: "
                f"expected {expected!r}, got {actual!r}"
            )


def force_recalc(output_path: Path) -> None:
    try:
        import xlwings as xw
    except ImportError:
        log.warning("xlwings not installed — open the file and press Ctrl+S once.")
        return
    log.info("Opening in Excel to force recalculation ...")
    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    app.screen_updating = False
    try:
        wb = app.books.open(str(output_path))
        app.calculate()
        wb.save()
        wb.close()
        log.info("   Recalculation done ✓")
    except Exception as e:
        log.warning("   Recalc failed: %s — open the file and press Ctrl+S once.", e)
    finally:
        app.screen_updating = True
        app.quit()


def run(config_path: Path) -> int:
    cfg = load_yaml(config_path)
    paths = build_pipeline_paths(cfg, config_path)
    workbooks = cfg.get("workbooks")
    if not isinstance(workbooks, dict):
        log.error("Config must define workbooks:")
        return 1

    spec = get_sheet_spec(workbooks, WB_NETWORK_RESOURCE, SHEET_NAME)
    hi = spec.header_row_1based - 1

    cur_file = resolve_workbook_path(paths.current_network_resource)
    prev_file = resolve_workbook_path(paths.previous_network_resource)
    if not cur_file:
        log.error("Current Network Resource workbook missing: %s", paths.current_network_resource)
        return 1
    if not prev_file:
        log.error("Previous Network Resource workbook missing: %s", paths.previous_network_resource)
        return 1

    log.info("Current  : %s", cur_file.name)
    log.info("Previous : %s", prev_file.name)

    rows_cur = load_grid(cur_file, SHEET_NAME)
    rows_prev = load_grid(prev_file, SHEET_NAME)
    if len(rows_cur) <= hi or len(rows_prev) <= hi:
        log.error("Sheet %r too short for header_row=%s", SHEET_NAME, spec.header_row_1based)
        return 1

    hdr_cur = list(rows_cur[hi])
    hdr_prev = list(rows_prev[hi])

    try:
        idx_cur = idx_map_from_headers(hdr_cur, spec.columns, NRR_COLUMN_KEYS)
        idx_prev = idx_map_from_headers(hdr_prev, spec.columns, PREV_KEYS)
    except (KeyError, ValueError) as e:
        log.error("%s", e)
        return 1

    cur_data = extract_table(rows_cur, spec.header_row_1based, idx_cur, NRR_COLUMN_KEYS)
    prev_data = extract_table(rows_prev, spec.header_row_1based, idx_prev, PREV_KEYS)

    log.info("Reading Fiber … %s current rows, %s previous rows", len(cur_data), len(prev_data))

    output_file = paths.fiber_computed_out
    output_file.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws_prev = wb.active
    ws_prev.title = "PrevWeek"

    pw_headers = ["Remarks", "FINALDWDMSpanValue", "Remark", "SpanCount"]
    for ci, h in enumerate(pw_headers, 1):
        cell = ws_prev.cell(row=1, column=ci, value=h)
        apply_header_style(cell, bg="085041")

    for ri, row in enumerate(prev_data, start=2):
        for ci, v in enumerate(row, 1):
            cell = ws_prev.cell(row=ri, column=ci, value=v)
            apply_data_style(cell, bg="E1F5EE" if ri % 2 == 0 else None)

    for ci in range(1, 5):
        ws_prev.column_dimensions[get_column_letter(ci)].width = 35
    ws_prev.sheet_state = "hidden"

    ws = wb.create_sheet(title=SHEET_NAME)
    all_headers = EXPECTED_GENERATED_HEADERS

    note_cell = ws.cell(
        row=1,
        column=32,
        value=(
            "AV: from AF when matched; if AF is N/A, AV uses AN (DWDM Span) as backup "
            "— Excel shows those AV cells in red."
        ),
    )
    note_cell.font = Font(italic=True, color="7F6000", name="Arial", size=9)
    note_cell.fill = _fill("FFEB9C")
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=32, end_row=1, end_column=49)
    ws.row_dimensions[1].height = 18

    for ci, h in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        if ci <= 21:
            apply_header_style(cell, bg="1F4E79")
        elif ci <= 31:
            apply_header_style(cell, bg="595959")
        else:
            apply_header_style(cell, bg="BF8F00")
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "A3"

    last_data_row = 2 + len(cur_data)
    for ri, row in enumerate(cur_data, start=3):
        alt = ri % 2 == 1
        for ci in range(1, 22):
            val = row[ci - 1] if (ci - 1) < len(row) else None
            apply_data_style(ws.cell(row=ri, column=ci, value=val), bg="DCE6F1" if alt else None)
        for ci in range(22, 32):
            val = row[ci - 1] if (ci - 1) < len(row) else None
            apply_data_style(
                ws.cell(row=ri, column=ci, value=val),
                bg="F2F2F2" if alt else "FAFAFA",
            )
        for col_letter, formula in formulas_for_row(ri).items():
            ci = column_index_from_string(col_letter)
            apply_data_style(
                ws.cell(row=ri, column=ci, value=formula),
                bg="FFF2CC" if alt else "FFFEF5",
            )

    if last_data_row >= 3:
        ws.conditional_formatting.add(
            f"AV3:AV{last_data_row}",
            FormulaRule(
                formula=['=$AF3="N/A"'],
                font=Font(color="9C0006", name="Arial", size=9, bold=True),
                fill=PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"),
            ),
        )

    width_map = {
        "A": 6, "B": 12, "C": 16, "D": 42, "E": 20, "F": 20,
        "G": 10, "H": 10, "I": 10, "J": 12, "K": 14, "L": 16,
        "M": 22, "N": 18, "O": 22, "P": 16, "Q": 12, "R": 12,
        "S": 14, "T": 12, "U": 28,
    }
    for col, w in width_map.items():
        ws.column_dimensions[col].width = w
    for ci in range(22, 32):
        ws.column_dimensions[get_column_letter(ci)].width = 8
    formula_widths = [28, 12, 14, 18, 12, 12, 16, 16, 26, 14, 16, 8, 8, 12, 12, 26, 28, 32]
    for i, w in enumerate(formula_widths):
        ws.column_dimensions[get_column_letter(32 + i)].width = w

    try:
        validate_generated_headers(ws)
    except ValueError as e:
        log.error("%s", e)
        return 1

    wb.save(str(output_file))
    log.info("Saved → %s", output_file.resolve())
    force_recalc(output_file)
    log.info("Done → %s", output_file.resolve())
    return 0


def main(argv: list[str] | None = None) -> int:
    setup_logging()
    p = argparse.ArgumentParser(description="Build Fiber computed workbook from ingest.yml")
    p.add_argument("--config", type=Path, required=True, help="Path to ingest YAML")
    args = p.parse_args(argv)
    cfg_path = args.config.resolve()
    if not cfg_path.is_file():
        log.error("Config not found: %s", cfg_path)
        return 1
    return run(cfg_path)


if __name__ == "__main__":
    sys.exit(main())
